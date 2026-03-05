# ui/source/source_page_energy.py
# Energy step UI and logic - mixin for SourcePage (split from main.py)

import sys
import os
import math
import threading
import calendar
import pandas as pd
import numpy as np
import customtkinter
import tkinter
from tkinter import filedialog
from ..toast import show_toast

from .constants import *
from .components import create_visualize_card, create_kpi_card
from .searchable_dropdown import SearchableDropdown

# Optional: source.panel_data
load_panels = get_manufacturers = get_curated_manufacturers = get_models_by_manufacturer = None
get_panel_params = get_panel_params_full = format_panel_summary = None
get_manufacturers_matching = get_combined_panel_results = parse_combined_selection = None
build_panel_index = None
try:
    from source.panel_data import (
        load_panels,
        build_panel_index,
        get_manufacturers,
        get_curated_manufacturers,
        get_models_by_manufacturer,
        get_panel_params,
        get_panel_params_full,
        format_panel_summary,
        get_manufacturers_matching,
        get_combined_panel_results,
        parse_combined_selection,
    )
except Exception as e:
    import traceback
    traceback.print_exc()
    print("Energy mixin (panel_data) import failed:", e)

# Optional: source.energy_models
TC_MODELS = {}
PPV_MODELS = {}
run_energy_calculation = None
aggregate_energy_hourly = None
prepare_energy_table_data = None
try:
    from source.energy_models import (
        TC_MODELS,
        PPV_MODELS,
        run_energy_calculation,
        aggregate_energy_hourly,
        prepare_energy_table_data,
    )
except Exception as e:
    import traceback
    traceback.print_exc()
    print("Energy mixin (energy_models) import failed:", e)

# Optional: source.fetch_fixed
fetch_fixed_mounting_hourly = None
try:
    from source.fetch_fixed import fetch_fixed_mounting_hourly
except Exception as e:
    import traceback
    traceback.print_exc()
    print("Energy mixin (fetch_fixed) import failed:", e)

# Optional: equation_renderer
TC_LATEX = {}
PPV_LATEX = {}
render_equations_to_image = None
format_constants_card = None
try:
    from .equation_renderer import (
        TC_LATEX,
        PPV_LATEX,
        render_equations_to_image,
        format_constants_card,
    )
except Exception:
    pass


def _derive_mounting_for_energy(cfg, source):
    """Derive mounting type for Energy step. Ninja: map tracking to mounting (like PVGIS)."""
    src = (source or "").upper()
    if src == "NINJA" and (cfg.get("ninja_mode") or "PV").upper() == "PV":
        tracking = (cfg.get("tracking") or "None").strip()
        return {"None": "fixed", "Single-axis": "inclined_axis", "Dual-axis": "two_axis"}.get(
            tracking, "fixed"
        )
    return (cfg.get("mounting_type") or cfg.get("mounting") or "fixed").strip().lower()




class SourcePageEnergyMixin:
    """Energy step: TC & Ppv models, panel, calculate, results, graphs, export."""

    def _save_energy_selections(self):
        """Save Energy dropdown selections before leaving Step 2 (e.g. when clicking Back)."""
        self._energy_saved_selections = None
        if not hasattr(self, "energy_manufacturer_combo") or not self.energy_manufacturer_combo.winfo_exists():
            return
        try:
            man = (self.energy_manufacturer_combo.get() or "").strip()
            panel = (self.energy_panel_combo.get() or "").strip() if hasattr(self, "energy_panel_combo") else ""
            tc = (self.energy_tc_combo.get() or "").strip() if hasattr(self, "energy_tc_combo") else ""
            ppv = (self.energy_ppv_combo.get() or "").strip() if hasattr(self, "energy_ppv_combo") else ""
            self._energy_saved_selections = {"man": man, "panel": panel, "tc": tc, "ppv": ppv}
        except Exception:
            self._energy_saved_selections = None

    def _restore_energy_selections(self):
        """Restore Energy dropdown selections after Step 2 is rebuilt (e.g. after Back then Next)."""
        saved = getattr(self, "_energy_saved_selections", None)
        if not saved or not hasattr(self, "energy_manufacturer_combo"):
            return
        try:
            man, panel, tc, ppv = saved.get("man", ""), saved.get("panel", ""), saved.get("tc", ""), saved.get("ppv", "")
            opts_mfr = getattr(self, "_energy_manufacturers", None) or []
            opts_tc = getattr(self, "_energy_tc_opts", None) or []
            opts_ppv = getattr(self, "_energy_ppv_opts", None) or []
            if man and opts_mfr and man in opts_mfr:
                self.energy_manufacturer_combo.set(man)
                self._on_energy_manufacturer_change(man)
            if panel and hasattr(self, "energy_panel_combo") and self.energy_panel_combo.winfo_exists():
                current_panel_opts = getattr(self.energy_panel_combo, "_values", [])
                if panel in current_panel_opts:
                    self.energy_panel_combo.set(panel)
                    self._on_energy_panel_change(panel)
            if tc and opts_tc and tc in opts_tc:
                self.energy_tc_combo.set(tc)
            if ppv and opts_ppv and ppv in opts_ppv:
                self.energy_ppv_combo.set(ppv)
            self._on_energy_model_change()
        except Exception:
            pass
        self._energy_saved_selections = None

    def _repopulate_energy_dropdowns(self):
        """Re-apply stored option lists to Energy dropdowns (e.g. after closing popup or on focus)."""
        try:
            opts_mfr = getattr(self, "_energy_manufacturers", None)
            opts_tc = getattr(self, "_energy_tc_opts", None)
            opts_ppv = getattr(self, "_energy_ppv_opts", None)
            if not opts_mfr and not opts_tc and not opts_ppv:
                return
            if hasattr(self, "energy_manufacturer_combo") and self.energy_manufacturer_combo.winfo_exists() and opts_mfr:
                cur = (self.energy_manufacturer_combo.get() or "").strip()
                self.energy_manufacturer_combo.set_values(opts_mfr)
                if cur and cur in opts_mfr:
                    self.energy_manufacturer_combo.set(cur)
            if hasattr(self, "energy_tc_combo") and self.energy_tc_combo.winfo_exists() and opts_tc:
                cur = (self.energy_tc_combo.get() or "").strip()
                self.energy_tc_combo.set_values(opts_tc)
                if cur and cur in opts_tc:
                    self.energy_tc_combo.set(cur)
            if hasattr(self, "energy_ppv_combo") and self.energy_ppv_combo.winfo_exists() and opts_ppv:
                cur = (self.energy_ppv_combo.get() or "").strip()
                self.energy_ppv_combo.set_values(opts_ppv)
                if cur and cur in opts_ppv:
                    self.energy_ppv_combo.set(cur)
            saved_panel = ""
            if hasattr(self, "energy_panel_combo") and self.energy_panel_combo.winfo_exists():
                saved_panel = (self.energy_panel_combo.get() or "").strip()
            if hasattr(self, "energy_manufacturer_combo") and self.energy_manufacturer_combo.winfo_exists():
                self._on_energy_manufacturer_change(self.energy_manufacturer_combo.get())
            if saved_panel and hasattr(self, "energy_panel_combo") and self.energy_panel_combo.winfo_exists():
                panel_opts = getattr(self.energy_panel_combo, "_values", [])
                if saved_panel in panel_opts:
                    self.energy_panel_combo.set(saved_panel)
        except Exception:
            pass

    def _build_step2_energy(self):
        """Step 2 Energy: TC & Ppv models, panel from SAM, results. Wired to Step 1 data."""
        if not hasattr(self, "_energy_result_df"):
            self._energy_result_df = None
        scroll = customtkinter.CTkScrollableFrame(
            self.content_inner, fg_color="transparent",
            scrollbar_button_color=THEME.border.gray,
            scrollbar_button_hover_color=THEME.border.medium,
        )
        scroll.pack(fill="both", expand=True)

        # ── Data from Step 1 — Info Chips ──
        cfg = self.import_config or {}
        source = cfg.get("source", "—")
        lat = cfg.get("latitude", "—")
        lon = cfg.get("longitude", "—")
        mount = _derive_mounting_for_energy(cfg, source) if source else cfg.get("mounting_type", "—")
        if isinstance(mount, str) and mount != "—":
            mount = mount.replace("_", " ").title()
        years = "—"
        if "start_year" in cfg and "end_year" in cfg:
            years = f"{cfg['start_year']}–{cfg['end_year']}"
        elif "year" in cfg:
            years = str(cfg["year"])
        try:
            api_result = cfg.get("api_result", {})
            hourly = api_result.get("hourly_data") if isinstance(api_result, dict) else None
            nrows = len(hourly) if hourly is not None else 0
        except Exception:
            nrows = 0

        # Info chips card
        chips_card = customtkinter.CTkFrame(
            scroll, fg_color=THEME.bg.gray_pale, corner_radius=RADIUS_MD,
            border_width=1, border_color=THEME.border.light,
        )
        chips_card.pack(fill="x", pady=(0, 16))
        chips_inner = customtkinter.CTkFrame(chips_card, fg_color="transparent")
        chips_inner.pack(fill="x", padx=16, pady=12)

        chip_data = [
            ("📡", "Source", str(source)),
            ("📍", "Location", f"{lat}, {lon}"),
            ("🔧", "Mounting", str(mount)),
            ("📅", "Period", str(years)),
        ]
        if nrows:
            chip_data.append(("📊", "Data Points", f"{fmt_num(nrows)} hourly"))

        for icon, label, value in chip_data:
            chip = customtkinter.CTkFrame(
                chips_inner, fg_color=BG_CARD, corner_radius=20,
                border_width=1, border_color=THEME.border.light,
            )
            chip.pack(side="left", padx=(0, 8), pady=2)
            customtkinter.CTkLabel(
                chip, text=f"{icon}  {label}: ", font=(FONT_FAMILY_TEXT, 11),
                text_color=TEXT_SECONDARY,
            ).pack(side="left", padx=(12, 0), pady=6)
            customtkinter.CTkLabel(
                chip, text=value, font=(FONT_FAMILY_TEXT, 11, "bold"),
                text_color=TEXT_PRIMARY,
            ).pack(side="left", padx=(0, 12), pady=6)

        # ── Panel (SAM) — Grouped Card ──
        self._energy_panels = load_panels() if load_panels else []
        self._energy_panel_index = build_panel_index(self._energy_panels) if build_panel_index else None
        idx = self._energy_panel_index
        manufacturers_base = (
            idx.get_curated_manufacturers(500) if idx else
            (get_curated_manufacturers(self._energy_panels, max_count=500) if get_curated_manufacturers else
             (get_manufacturers(self._energy_panels) if get_manufacturers else []))
        )
        manufacturers = ["— Select —"] + manufacturers_base

        def _manufacturer_filter(query: str):
            """Custom filter: uses cached index for fast lookups."""
            idx = getattr(self, "_energy_panel_index", None)
            if not query or not (query or "").strip() or (query or "").strip().lower() in ("— select —",):
                base = idx.get_curated_manufacturers(500) if idx else (
                    get_curated_manufacturers(self._energy_panels, max_count=500) if get_curated_manufacturers else
                    (get_manufacturers(self._energy_panels) if get_manufacturers else [])
                )
                return ["— Select —"] + base
            if idx:
                mfrs = idx.get_manufacturers_matching(query)
                combined = idx.get_combined_panel_results(query)
            else:
                mfrs = get_manufacturers_matching(self._energy_panels, query) if get_manufacturers_matching else []
                combined = get_combined_panel_results(self._energy_panels, query) if get_combined_panel_results else []
            return ["— Select —"] + mfrs + combined
        panel_card = customtkinter.CTkFrame(
            scroll, fg_color=BG_CARD, corner_radius=RADIUS_MD,
            border_width=1, border_color=THEME.border.light,
        )
        panel_card.pack(fill="x", pady=(0, 12))

        # Section header
        panel_header = customtkinter.CTkFrame(panel_card, fg_color="transparent")
        panel_header.pack(fill="x", padx=20, pady=(16, 4))
        customtkinter.CTkLabel(
            panel_header, text="⚡", font=(FONT_FAMILY_DISPLAY, 18),
        ).pack(side="left", padx=(0, 8))
        customtkinter.CTkLabel(
            panel_header, text="Panel (SAM)",
            font=(FONT_FAMILY_TEXT, 15, "bold"), text_color=TEXT_PRIMARY,
        ).pack(side="left")
        customtkinter.CTkLabel(
            panel_header, text="Select manufacturer and panel type",
            font=(FONT_FAMILY_TEXT, 11), text_color=TEXT_MUTED,
        ).pack(side="left", padx=(12, 0))

        # Divider
        customtkinter.CTkFrame(panel_card, fg_color=THEME.border.light, height=1).pack(fill="x", padx=20, pady=(8, 0))

        # Panel fields — vertical label-above-field layout
        panel_fields = customtkinter.CTkFrame(panel_card, fg_color="transparent")
        panel_fields.pack(fill="x", padx=20, pady=(12, 16))
        panel_fields.columnconfigure(0, weight=1)
        panel_fields.columnconfigure(1, weight=1)

        # Manufacturer
        mfr_frame = customtkinter.CTkFrame(panel_fields, fg_color="transparent")
        mfr_frame.grid(row=0, column=0, sticky="ew", padx=(0, 12))
        customtkinter.CTkLabel(
            mfr_frame, text="Manufacturer", font=(FONT_FAMILY_TEXT, 11, "bold"),
            text_color=TEXT_SECONDARY,
        ).pack(anchor="w", pady=(0, 4))
        self.energy_manufacturer_combo = SearchableDropdown(
            mfr_frame,
            values=manufacturers,
            width=280,
            placeholder="— Select —",
            command=self._on_energy_manufacturer_change,
            fg_color=THEME.bg.gray_pale if hasattr(THEME, "bg") else None,
            combo_style=True,
            custom_filter=_manufacturer_filter,
            max_display_options=150,
        )
        self.energy_manufacturer_combo.pack(anchor="w")

        # Panel type
        def _panel_type_filter(query: str):
            """Filter with relevance: exact match → startswith → contains."""
            combo = getattr(self, "energy_panel_combo", None)
            if combo is None:
                return ["— Select —"]
            vals = getattr(combo, "_values", []) or []
            sentinel = "— Select —"
            if not query or not (query or "").strip() or (query or "").strip().lower() in ("— select —",):
                return vals
            q = (query or "").strip().lower()
            rest = [v for v in vals if v != sentinel]
            exact = [v for v in rest if v.lower() == q]
            startswith = sorted([v for v in rest if v.lower().startswith(q) and v not in exact])
            contains = sorted([v for v in rest if q in v.lower() and v not in exact and v not in startswith])
            result = exact + startswith + contains
            return [sentinel] + result if sentinel in vals else result

        ptype_frame = customtkinter.CTkFrame(panel_fields, fg_color="transparent")
        ptype_frame.grid(row=0, column=1, sticky="ew", padx=(12, 0))
        customtkinter.CTkLabel(
            ptype_frame, text="Panel Type", font=(FONT_FAMILY_TEXT, 11, "bold"),
            text_color=TEXT_SECONDARY,
        ).pack(anchor="w", pady=(0, 4))
        self.energy_panel_combo = SearchableDropdown(
            ptype_frame,
            values=["— Select —"],
            width=280,
            placeholder="Type to search (e.g. JKM5, 370, 72HL)",
            command=self._on_energy_panel_change,
            fg_color=THEME.bg.gray_pale if hasattr(THEME, "bg") else None,
            empty_message="Select a manufacturer first to see panel models.",
            combo_style=True,
            custom_filter=_panel_type_filter,
            max_display_options=150,
        )
        self.energy_panel_combo.pack(anchor="w")

        self.energy_panel_summary_label = customtkinter.CTkLabel(
            panel_card, text="",
            font=(FONT_FAMILY_TEXT, 11), text_color=TEXT_MUTED,
        )
        self.energy_panel_summary_label.pack(anchor="w", padx=20, pady=(0, 12))

        # ── TC model & Ppv model — Grouped Card ──
        tc_opts = ["— Select TC model —"] + list(TC_MODELS.keys())
        ppv_opts = ["— Select Ppv model —"] + list(PPV_MODELS.keys())
        self._energy_manufacturers = manufacturers
        self._energy_tc_opts = tc_opts
        self._energy_ppv_opts = ppv_opts

        model_card = customtkinter.CTkFrame(
            scroll, fg_color=BG_CARD, corner_radius=RADIUS_MD,
            border_width=1, border_color=THEME.border.light,
        )
        model_card.pack(fill="x", pady=(0, 12))

        # Section header
        model_header = customtkinter.CTkFrame(model_card, fg_color="transparent")
        model_header.pack(fill="x", padx=20, pady=(16, 4))
        customtkinter.CTkLabel(
            model_header, text="📐", font=(FONT_FAMILY_DISPLAY, 18),
        ).pack(side="left", padx=(0, 8))
        customtkinter.CTkLabel(
            model_header, text="Models",
            font=(FONT_FAMILY_TEXT, 15, "bold"), text_color=TEXT_PRIMARY,
        ).pack(side="left")
        customtkinter.CTkLabel(
            model_header, text="Cell temperature & PV power models",
            font=(FONT_FAMILY_TEXT, 11), text_color=TEXT_MUTED,
        ).pack(side="left", padx=(12, 0))

        # Divider
        customtkinter.CTkFrame(model_card, fg_color=THEME.border.light, height=1).pack(fill="x", padx=20, pady=(8, 0))

        # Model fields — vertical label-above-field layout
        model_fields = customtkinter.CTkFrame(model_card, fg_color="transparent")
        model_fields.pack(fill="x", padx=20, pady=(12, 16))
        model_fields.columnconfigure(0, weight=1)
        model_fields.columnconfigure(1, weight=1)

        # TC model
        tc_frame = customtkinter.CTkFrame(model_fields, fg_color="transparent")
        tc_frame.grid(row=0, column=0, sticky="ew", padx=(0, 12))
        customtkinter.CTkLabel(
            tc_frame, text="TC Model (Cell Temperature)", font=(FONT_FAMILY_TEXT, 11, "bold"),
            text_color=TEXT_SECONDARY,
        ).pack(anchor="w", pady=(0, 4))
        self.energy_tc_combo = SearchableDropdown(
            tc_frame,
            values=tc_opts,
            width=280,
            placeholder="— Select TC model —",
            command=lambda v: self._on_energy_model_change(),
            fg_color=THEME.bg.gray_pale if hasattr(THEME, "bg") else None,
            combo_style=True,
        )
        self.energy_tc_combo.set("— Select TC model —")
        self.energy_tc_combo.pack(anchor="w")

        # Ppv model
        ppv_frame = customtkinter.CTkFrame(model_fields, fg_color="transparent")
        ppv_frame.grid(row=0, column=1, sticky="ew", padx=(12, 0))
        customtkinter.CTkLabel(
            ppv_frame, text="Ppv Model (PV Power)", font=(FONT_FAMILY_TEXT, 11, "bold"),
            text_color=TEXT_SECONDARY,
        ).pack(anchor="w", pady=(0, 4))
        self.energy_ppv_combo = SearchableDropdown(
            ppv_frame,
            values=ppv_opts,
            width=280,
            placeholder="— Select Ppv model —",
            command=lambda v: self._on_energy_model_change(),
            fg_color=THEME.bg.gray_pale if hasattr(THEME, "bg") else None,
            combo_style=True,
        )
        self.energy_ppv_combo.set("— Select Ppv model —")
        self.energy_ppv_combo.pack(anchor="w")

        # Optional: repopulate lists when dropdown gets focus (e.g. after Calculate or popup close)
        for combo_attr in ("energy_manufacturer_combo", "energy_tc_combo", "energy_ppv_combo"):
            combo = getattr(self, combo_attr, None)
            if combo is not None and hasattr(combo, "entry"):
                try:
                    combo.entry.bind("<FocusIn>", lambda e, s=self: s._repopulate_energy_dropdowns())
                except Exception:
                    pass

        # ── Equation display — Enhanced cards ──
        eq_section_card = customtkinter.CTkFrame(
            scroll, fg_color=BG_CARD, corner_radius=RADIUS_MD,
            border_width=1, border_color=THEME.border.light,
        )
        eq_section_card.pack(fill="x", pady=(0, 12))

        eq_header = customtkinter.CTkFrame(eq_section_card, fg_color="transparent")
        eq_header.pack(fill="x", padx=20, pady=(16, 4))
        customtkinter.CTkLabel(
            eq_header, text="📝", font=(FONT_FAMILY_DISPLAY, 18),
        ).pack(side="left", padx=(0, 8))
        customtkinter.CTkLabel(
            eq_header, text="Equations & Constants",
            font=(FONT_FAMILY_TEXT, 15, "bold"), text_color=TEXT_PRIMARY,
        ).pack(side="left")

        customtkinter.CTkFrame(eq_section_card, fg_color=THEME.border.light, height=1).pack(fill="x", padx=20, pady=(8, 0))

        cards_row = customtkinter.CTkFrame(eq_section_card, fg_color="transparent")
        cards_row.pack(fill="x", padx=16, pady=(12, 16))
        cards_row.columnconfigure(0, weight=1)
        cards_row.columnconfigure(1, weight=1)

        card_style = dict(
            fg_color=THEME.bg.gray_pale,
            corner_radius=10,
            border_width=1,
            border_color=THEME.border.light,
        )
        self.energy_eq_card = customtkinter.CTkFrame(cards_row, **card_style)
        self.energy_eq_card.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        self.energy_const_card = customtkinter.CTkFrame(cards_row, **card_style)
        self.energy_const_card.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

        # Equation card badge
        eq_badge = customtkinter.CTkFrame(self.energy_eq_card, fg_color=THEME.bg.gray_pale, corner_radius=6)
        eq_badge.pack(anchor="w", padx=12, pady=(12, 4))
        customtkinter.CTkLabel(
            eq_badge, text=" Equations ", font=(FONT_FAMILY_TEXT, 10, "bold"),
            text_color=TEXT_PRIMARY,
        ).pack(padx=6, pady=2)

        self.energy_eq_inner = customtkinter.CTkFrame(
            self.energy_eq_card, fg_color="transparent",
            border_width=1, border_color=THEME.border.light, corner_radius=8,
        )
        self.energy_eq_inner.pack(fill="both", expand=True, padx=12, pady=(4, 12))
        self.energy_eq_inner.configure(height=120)
        self.energy_equation_label = customtkinter.CTkLabel(
            self.energy_eq_inner,
            text="Select a TC model and a Ppv model to show equations.",
            font=(FONT_FAMILY_TEXT, 12), text_color=TEXT_SECONDARY,
            wraplength=320, justify="left",
        )
        self.energy_equation_label.pack(anchor="w", padx=8, pady=8)
        self._energy_eq_image_ref = None

        # Constants card badge
        const_badge = customtkinter.CTkFrame(self.energy_const_card, fg_color=THEME.bg.gray_pale, corner_radius=6)
        const_badge.pack(anchor="w", padx=12, pady=(12, 4))
        customtkinter.CTkLabel(
            const_badge, text=" Constants ", font=(FONT_FAMILY_TEXT, 10, "bold"),
            text_color=TEXT_PRIMARY,
        ).pack(padx=6, pady=2)

        self.energy_constants_label = customtkinter.CTkLabel(
            self.energy_const_card,
            text="Select a panel and models to see constants.",
            font=("Consolas", 13), text_color=TEXT_SECONDARY,
            wraplength=320, justify="left",
        )
        self.energy_constants_label.pack(anchor="nw", padx=16, pady=(4, 14))


        # ── Calculate — Action Card ──
        action_card = customtkinter.CTkFrame(
            scroll, fg_color=BG_CARD, corner_radius=RADIUS_MD,
            border_width=1, border_color=THEME.border.light,
        )
        action_card.pack(fill="x", pady=(0, 8))
        btn_row = customtkinter.CTkFrame(action_card, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=16)

        self.energy_calculate_btn = customtkinter.CTkButton(
            btn_row, text="⚡  Calculate TC(f) & Ppv(f)",
            command=self._on_energy_calculate,
            fg_color=PRIMARY_BLUE, hover_color=PRIMARY_HOVER, text_color="#FFFFFF",
            corner_radius=10, width=240, height=42, font=(FONT_FAMILY_TEXT, 13, "bold"),
        )
        self.energy_calculate_btn.pack(side="left", padx=(0, 12))
        self.energy_visualize_export_btn = customtkinter.CTkButton(
            btn_row, text="📊  Visualize & Export",
            command=self._open_energy_visualize_export_popup,
            fg_color=THEME.bg.gray_pale, border_width=1, border_color=THEME.border.light,
            hover_color=THEME.bg.hover, text_color=TEXT_SECONDARY,
            corner_radius=10, width=170, height=42, font=(FONT_FAMILY_TEXT, 12, "bold"),
            state="disabled",
        )
        self.energy_visualize_export_btn.pack(side="left")
        self.energy_status_label = customtkinter.CTkLabel(
            btn_row, text="",
            font=(FONT_FAMILY_TEXT, 12), text_color=TEXT_SECONDARY,
        )
        self.energy_status_label.pack(side="left", padx=(16, 0))

        self._on_energy_model_change()
        self._restore_energy_selections()

    def _on_energy_manufacturer_change(self, choice=None):
        """Populate Panel type dropdown from selected manufacturer. Handle combined 'Manufacturer — Model' selection."""
        if not get_models_by_manufacturer or not hasattr(self, "energy_panel_combo"):
            return
        raw = (choice or self.energy_manufacturer_combo.get() or "").strip()
        if not raw or raw == "— Select —":
            self.energy_panel_combo.set_values(["— Select —"])
            self.energy_panel_combo.set("— Select —")
            self.energy_panel_combo.set_empty_message("Select a manufacturer first to see panel models.")
            if hasattr(self, "energy_panel_summary_label"):
                self.energy_panel_summary_label.configure(text="")
            return
        man, model_from_combined = parse_combined_selection(raw) if parse_combined_selection else (raw, None)
        if not man or man == "— Select —":
            return
        if model_from_combined and hasattr(self, "energy_manufacturer_combo"):
            self.energy_manufacturer_combo.set(man)
        idx = getattr(self, "_energy_panel_index", None)
        models = idx.get_models_by_manufacturer(man) if idx else get_models_by_manufacturer(self._energy_panels, man)
        # Deduplicate and sort for predictable UX
        model_names = sorted(dict.fromkeys(p.get("model", "") for p in models if p.get("model")))
        self.energy_panel_combo.set_values(["— Select —"] + model_names)
        if model_from_combined and model_from_combined in model_names:
            self.energy_panel_combo.set(model_from_combined)
            if hasattr(self, "_on_energy_panel_change"):
                self._on_energy_panel_change(model_from_combined)
        else:
            self.energy_panel_combo.set("— Select —")
        if not model_names:
            self.energy_panel_combo.set_empty_message("No models found. Try another manufacturer or check your data.")
            if hasattr(self, "energy_panel_summary_label"):
                self.energy_panel_summary_label.configure(
                    text="No panel models found for this manufacturer.",
                    text_color=TEXT_MUTED,
                )
        else:
            self.energy_panel_combo.set_empty_message("")
            if hasattr(self, "energy_panel_summary_label") and not model_from_combined:
                self.energy_panel_summary_label.configure(text="")

    def _on_energy_panel_change(self, choice=None):
        """Update panel summary line and constants card when a panel type is selected."""
        if not get_panel_params or not hasattr(self, "energy_panel_summary_label"):
            return
        man = (self.energy_manufacturer_combo.get() or "").strip() if hasattr(self, "energy_manufacturer_combo") else ""
        model = (choice or (self.energy_panel_combo.get() if hasattr(self, "energy_panel_combo") else "")).strip()
        if not man or man == "— Select —" or not model or model == "— Select —":
            self.energy_panel_summary_label.configure(text="")
        else:
            idx = getattr(self, "_energy_panel_index", None)
            panel = (idx.get_panel_params(man, model) if idx else get_panel_params(self._energy_panels, man, model))
            if format_panel_summary:
                self.energy_panel_summary_label.configure(text=format_panel_summary(panel) if panel else "")
        if hasattr(self, "_on_energy_model_change"):
            self._on_energy_model_change()

    def _on_energy_model_change(self):
        """Update equation card (LaTeX image) and constants card from selected TC, Ppv, and panel."""
        if not hasattr(self, "energy_eq_inner"):
            return
        tc_id = (self.energy_tc_combo.get() or "").strip()
        ppv_id = (self.energy_ppv_combo.get() or "").strip()
        has_tc = tc_id and tc_id != "— Select TC model —" and tc_id in TC_MODELS
        has_ppv = ppv_id and ppv_id != "— Select Ppv model —" and ppv_id in PPV_MODELS

        # Clear equation card inner and show either LaTeX image or text placeholder
        for w in self.energy_eq_inner.winfo_children():
            w.destroy()
        self._energy_eq_image_ref = None

        if has_tc and has_ppv and render_equations_to_image and TC_LATEX and PPV_LATEX:
            tc_tex = TC_LATEX.get(tc_id)
            ppv_tex = PPV_LATEX.get(ppv_id)
            bg = getattr(getattr(THEME, "bg", None), "gray_pale", "#FAFAFA") or "#FAFAFA"
            result = render_equations_to_image(
                tc_tex, ppv_tex,
                width_inches=6.5, dpi=150, bg_color=bg, fontsize=18,
            )
            if result:
                pil_img, (w, h) = result
                try:
                    max_w, max_h = 450, 320
                    scale = min(max_w / w, max_h / h, 1.0)
                    sz = (int(w * scale), int(h * scale))
                    ctk_img = customtkinter.CTkImage(
                        light_image=pil_img,
                        dark_image=pil_img,
                        size=sz,
                    )
                    self._energy_eq_image_ref = ctk_img
                    lbl = customtkinter.CTkLabel(
                        self.energy_eq_inner, image=ctk_img, text="",
                        fg_color="transparent",
                    )
                    lbl.pack(anchor="w")
                except Exception:
                    pass
        if not self.energy_eq_inner.winfo_children():
            customtkinter.CTkLabel(
                self.energy_eq_inner,
                text="Select a TC model and a Ppv model to show equations." if not (has_tc and has_ppv)
                else "Equation rendering unavailable.",
                font=("Segoe UI", 12), text_color=TEXT_SECONDARY,
                wraplength=320, justify="left",
            ).pack(anchor="w")

        # Constants card: panel values for selected models
        tc_const = list(TC_MODELS.get(tc_id, {}).get("constants", [])) if has_tc else []
        ppv_const = list(PPV_MODELS.get(ppv_id, {}).get("constants", [])) if has_ppv else []
        panel = None
        peak = 1.0
        if get_panel_params_full and hasattr(self, "_energy_panels"):
            man = (self.energy_manufacturer_combo.get() or "").strip()
            mdl = (self.energy_panel_combo.get() or "").strip()
            if man and man != "— Select —" and mdl and mdl != "— Select —":
                panel = get_panel_params_full(self._energy_panels, man, mdl)
            cfg = self.import_config or {}
            peak = float(cfg.get("peak_power_kwp") or cfg.get("peak_power") or cfg.get("capacity") or 1.0)
        constants_text = format_constants_card(tc_const, ppv_const, panel, peak) if format_constants_card else "Select a panel and models."
        if hasattr(self, "energy_constants_label"):
            self.energy_constants_label.configure(text=constants_text)

    def _on_energy_calculate_done(self, result_df, err, import_hourly_df=None):
        """Called on main thread when background calculation finishes. Enable button, show result."""
        if hasattr(self, "energy_calculate_btn") and self.energy_calculate_btn.winfo_exists():
            self.energy_calculate_btn.configure(state="normal")
        if result_df is None or (hasattr(result_df, "empty") and result_df.empty):
            self.energy_status_label.configure(text=err or "No results.", text_color=ERROR_RED)
            try:
                self._repopulate_energy_dropdowns()
            except Exception:
                pass
            return
        # Merge import data, fill NaN→0 for power/energy, ensure physical sense
        if prepare_energy_table_data:
            result_df = prepare_energy_table_data(result_df, import_hourly_df)
        self._energy_result_df = result_df
        if err:
            self.energy_status_label.configure(text=err, text_color=INFO_BLUE)
        else:
            self.energy_status_label.configure(text=f"Done: {fmt_num(len(result_df))} rows.", text_color=SUCCESS_GREEN)
        self._update_energy_visualize_export_button_state()
        try:
            self._repopulate_energy_dropdowns()
        except Exception:
            pass

    def _on_energy_calculate(self):
        """Run TC(f) and Ppv(f) in a background thread; show progress text until done."""
        if not run_energy_calculation or not get_panel_params:
            self.energy_status_label.configure(text="Energy models not available.", text_color=ERROR_RED)
            return
        man = self.energy_manufacturer_combo.get().strip()
        model_name = self.energy_panel_combo.get().strip()
        if not man or man == "— Select —" or not model_name or model_name == "— Select —":
            self.energy_status_label.configure(text="Select Manufacturer and Panel type.", text_color=ERROR_RED)
            return
        panel = get_panel_params(self._energy_panels, man, model_name)
        if not panel:
            self.energy_status_label.configure(text="Panel not found.", text_color=ERROR_RED)
            return
        tc_id = self.energy_tc_combo.get().strip()
        ppv_id = self.energy_ppv_combo.get().strip()
        if not tc_id or tc_id == "— Select TC model —" or not ppv_id or ppv_id == "— Select Ppv model —":
            self.energy_status_label.configure(text="Select TC and Ppv models.", text_color=ERROR_RED)
            return
        cfg = self.import_config or {}
        api_result = cfg.get("api_result")
        if not api_result or not isinstance(api_result, dict):
            self.energy_status_label.configure(text="No Step 1 data. Import data in Step 1 first.", text_color=ERROR_RED)
            return
        hourly_df = api_result.get("hourly_data")
        if hourly_df is None or (hasattr(hourly_df, "empty") and hourly_df.empty):
            self.energy_status_label.configure(text="No hourly data from Step 1.", text_color=ERROR_RED)
            return

        self.energy_status_label.configure(text="Calculating TC(f) & Ppv(f)...", text_color=TEXT_PRIMARY)
        if hasattr(self, "energy_calculate_btn") and self.energy_calculate_btn.winfo_exists():
            self.energy_calculate_btn.configure(state="disabled")

        def run_in_thread():
            try:
                source = cfg.get("source", "PVGIS")
                peak = float(cfg.get("peak_power_kwp") or cfg.get("peak_power") or cfg.get("capacity") or 1.0)
                mounting = _derive_mounting_for_energy(cfg, source)
                hourly_fixed_df = None
                if mounting != "fixed" and fetch_fixed_mounting_hourly:
                    hourly_fixed_df, fixed_err = fetch_fixed_mounting_hourly(cfg)
                    if fixed_err and hourly_fixed_df is None:
                        self.after(0, lambda: self._on_energy_calculate_done(None, fixed_err, None))
                        return
                result_df, err = run_energy_calculation(
                    hourly_df, source, panel, peak,
                    tc_model_id=tc_id, ppv_model_id=ppv_id,
                    mounting_type=mounting, hourly_fixed_df=hourly_fixed_df,
                )
                self.after(0, lambda: self._on_energy_calculate_done(result_df, err, hourly_df))
            except Exception as e:
                self.after(0, lambda: self._on_energy_calculate_done(None, str(e), None))

        threading.Thread(target=run_in_thread, daemon=True).start()

    def _update_energy_visualize_export_button_state(self):
        """Enable Visualize & Export button only after successful calculation. Blue when active, grey when disabled."""
        if not hasattr(self, "energy_visualize_export_btn") or not self.energy_visualize_export_btn.winfo_exists():
            return
        df = getattr(self, "_energy_result_df", None)
        has_results = df is not None and (not hasattr(df, "empty") or not df.empty)
        if has_results:
            self.energy_visualize_export_btn.configure(
                state="normal",
                fg_color=PRIMARY_BLUE,
                hover_color=PRIMARY_HOVER,
                text_color="#FFFFFF",
            )
        else:
            self.energy_visualize_export_btn.configure(
                state="disabled",
                fg_color=THEME.bg.gray_pale,
                border_color=THEME.border.light,
                text_color=THEME.text.muted,
            )

    def _open_energy_visualize_export_popup(self):
        """Open Energy Visualize & Export popup with Metadata, Result table, Graphs, Export tabs."""
        df = getattr(self, "_energy_result_df", None)
        if df is None or (hasattr(df, "empty") and df.empty):
            show_toast(self, "No results to display. Run Calculate first.", type="info")
            return
        popup = customtkinter.CTkToplevel(self)
        popup.title("Visualize & Export — Energy Results")
        popup.geometry("950x650")
        popup.configure(fg_color=THEME.bg.main)
        try:
            popup.transient(self.winfo_toplevel())
        except Exception:
            pass

        def on_popup_close():
            if hasattr(self, "_visualize_export_modal_root") and self._visualize_export_modal_root is popup:
                self._visualize_export_modal_root = None
            try:
                self._repopulate_energy_dropdowns()
            except Exception:
                pass
            try:
                popup.destroy()
            except Exception:
                pass

        self._visualize_export_modal_root = popup
        try:
            popup.protocol("WM_DELETE_WINDOW", on_popup_close)
        except Exception:
            pass

        inner = customtkinter.CTkFrame(popup, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=24, pady=24)

        tabview = customtkinter.CTkTabview(
            inner, fg_color="transparent",
            segmented_button_fg_color=THEME.bg.hover,
            segmented_button_selected_color=THEME.primary.blue,
            segmented_button_selected_hover_color=THEME.primary.blue_hover,
            segmented_button_unselected_color=THEME.bg.hover,
            segmented_button_unselected_hover_color=THEME.border.light,
            text_color="#FFFFFF",
            text_color_disabled=THEME.text.secondary,
            corner_radius=10,
        )
        tabview.pack(fill="both", expand=True, padx=8, pady=(8, 4))

        tab_meta = tabview.add("Metadata")
        tab_table = tabview.add("Result table")
        tab_graphs = tabview.add("Graphs")
        tab_export = tabview.add("Export")

        self._build_energy_metadata_tab(tab_meta)
        self._build_energy_result_table_tab(tab_table)
        self._build_energy_graphs_tab(tab_graphs)
        self._build_energy_export_tab(tab_export)

    def _build_energy_metadata_tab(self, parent):
        """Build Metadata tab: Step 1-style icon rows + KPI cards for CF, panels, power."""
        scroll = customtkinter.CTkScrollableFrame(
            parent, fg_color="transparent",
            scrollbar_button_color=THEME.border.gray,
            scrollbar_button_hover_color=THEME.border.medium,
        )
        scroll.pack(fill="both", expand=True, padx=30, pady=30)

        customtkinter.CTkLabel(
            scroll, text="Energy Calculation Metadata",
            font=(FONT_FAMILY_DISPLAY, 18, "bold"), text_color=TEXT_PRIMARY,
        ).pack(anchor="w", pady=(0, 20))

        # ── Badge ──
        badge = customtkinter.CTkLabel(
            scroll, text="⚡ Energy Calculation",
            font=(FONT_FAMILY_TEXT, 14, "bold"),
            text_color=THEME.status.success_alt if hasattr(THEME.status, "success_alt") else SUCCESS_GREEN,
            fg_color=THEME.bg.gray_light if hasattr(THEME.bg, "gray_light") else THEME.bg.gray_lighter,
            corner_radius=8, padx=15, pady=8,
        )
        badge.pack(anchor="w", pady=(0, 20))

        cfg = self.import_config or {}
        source = cfg.get("source", "N/A")
        lat = cfg.get("latitude", "N/A")
        lon = cfg.get("longitude", "N/A")
        peak_raw = cfg.get("peak_power_kwp") or cfg.get("peak_power") or cfg.get("capacity_kw") or cfg.get("capacity")
        try:
            peak_kw = float(peak_raw) if peak_raw is not None else 1.0
        except (TypeError, ValueError):
            peak_kw = 1.0
        peak = f"{peak_kw:.2f}" if peak_raw is not None else "N/A"
        man = (self.energy_manufacturer_combo.get() or "").strip() if hasattr(self, "energy_manufacturer_combo") else "N/A"
        mdl = (self.energy_panel_combo.get() or "").strip() if hasattr(self, "energy_panel_combo") else "N/A"
        tc_id = (self.energy_tc_combo.get() or "").strip() if hasattr(self, "energy_tc_combo") else "N/A"
        ppv_id = (self.energy_ppv_combo.get() or "").strip() if hasattr(self, "energy_ppv_combo") else "N/A"
        df = getattr(self, "_energy_result_df", None)
        nrows = len(df) if df is not None and (not hasattr(df, "empty") or not df.empty) else 0

        # ── Details card with icon rows (Step 1 style) ──
        details = customtkinter.CTkFrame(scroll, fg_color=THEME.bg.gray_lighter, corner_radius=10)
        details.pack(fill="x", pady=10)

        def add_config_row(frame, icon, label, value):
            row = customtkinter.CTkFrame(frame, fg_color="transparent")
            row.pack(fill="x", padx=20, pady=8)
            customtkinter.CTkLabel(row, text=icon, font=(FONT_FAMILY_TEXT, 16), width=30).pack(side="left")
            customtkinter.CTkLabel(row, text=label, font=(FONT_FAMILY_TEXT, 12), text_color=TEXT_SECONDARY, width=120).pack(side="left")
            customtkinter.CTkLabel(row, text=str(value), font=(FONT_FAMILY_TEXT, 12, "bold"), text_color=TEXT_PRIMARY).pack(side="left", padx=(10, 0))

        add_config_row(details, "📊", "Data Source:", source)
        add_config_row(details, "📍", "Location:", f"Lat {lat}°, Lon {lon}°")
        add_config_row(details, "⚡", "Peak Power:", f"{peak} kWp" if peak != "N/A" else peak)
        add_config_row(details, "🔋", "Panel:", f"{man} — {mdl}")
        add_config_row(details, "🌡️", "TC Model:", tc_id)
        add_config_row(details, "📐", "Ppv Model:", ppv_id)
        add_config_row(details, "📈", "Result Rows:", fmt_num(nrows))

        # ── Compute KPI values ──
        n_panels_val = "N/A"
        n_panels_unit = "panels"
        real_power_val = "N/A"
        real_power_unit = "kWp"
        cf_val = "N/A"
        cf_unit = "capacity factor"
        if get_panel_params and peak_kw > 0 and man and man != "N/A" and mdl and mdl != "N/A":
            panel = get_panel_params(getattr(self, "_energy_panels", []), man, mdl)
            if panel:
                pmax = float(panel.get("Pmax", 300) or 300)
                if pmax > 0:
                    n_panels = math.ceil(peak_kw * 1000.0 / pmax)
                    n_panels_val = fmt_num(n_panels)
                    real_power_w = n_panels * pmax
                    real_power_val = f"{fmt_num(real_power_w / 1000, 2)}"
        if df is not None and (not hasattr(df, "empty") or not df.empty) and "Ppv(f)" in getattr(df, "columns", []):
            pnom_w = peak_kw * 1000.0
            n_hours = len(df)
            if pnom_w > 0 and n_hours > 0:
                try:
                    total_wh = float(df["Ppv(f)"].sum())
                    cf = total_wh / (pnom_w * n_hours)
                    cf_val = f"{cf * 100:.2f}%"
                except (TypeError, ValueError):
                    pass

        # ── KPI Cards row ──
        kpi_frame = customtkinter.CTkFrame(scroll, fg_color="transparent")
        kpi_frame.pack(fill="x", pady=(20, 0))
        kpi_frame.grid_columnconfigure((0, 1, 2), weight=1, uniform="kpi")

        create_kpi_card(kpi_frame, "Number of Panels", n_panels_val, n_panels_unit, row=0, column=0)
        create_kpi_card(kpi_frame, "Real Power Plant", real_power_val, real_power_unit, row=0, column=1)
        create_kpi_card(kpi_frame, "Capacity Factor", cf_val, cf_unit, row=0, column=2)

    def _get_peak_power_kw(self):
        """Peak power (kW) from import config, default 1.0."""
        cfg = getattr(self, "import_config", None) or {}
        raw = cfg.get("peak_power_kwp") or cfg.get("peak_power") or cfg.get("capacity_kw") or cfg.get("capacity")
        try:
            return float(raw) if raw is not None else 1.0
        except (TypeError, ValueError):
            return 1.0

    def _add_cf_columns(self, agg_df, rule):
        """Add CF(f), CF(x) to aggregated df. rule: 'D'|'ME'|'YE'."""
        from source.energy_models import add_cf_columns
        peak_kw = self._get_peak_power_kw()
        return add_cf_columns(agg_df, rule, peak_kw)

    def _is_tmy_import(self):
        """True when import is PVGIS TMY."""
        cfg = getattr(self, "import_config", None) or {}
        return cfg.get("source") == "PVGIS" and cfg.get("pvgis_mode") == "TMY"

    def _drop_empty_aggregate_rows(self, df):
        """Drop rows where all key energy cols are 0/NaN. Used for TMY to remove empty aggregated periods."""
        if df is None or df.empty:
            return df
        energy_cols = [c for c in ("Epv(f)", "Epv(x)", "Esource(f)", "Esource(x)") if c in df.columns]
        if not energy_cols:
            return df
        vals = df[energy_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
        has_data = (vals > 0).any(axis=1)
        return df.loc[has_data].reset_index(drop=True)

    def _compute_energy_aggregate(self, hourly_df, rule, add_month_year=False):
        """Aggregate energy hourly. Power cols: SUM; TC, Kt: MEAN. rule: 'D'|'ME'|'YE'."""
        if aggregate_energy_hourly is None:
            return pd.DataFrame()
        agg = aggregate_energy_hourly(hourly_df, rule, add_month_year)
        return self._add_cf_columns(agg, rule)

    def _compute_energy_daily_view(self, hourly_df):
        return self._compute_energy_aggregate(hourly_df, "D", add_month_year=False)

    def _compute_energy_monthly_view(self, hourly_df):
        return self._compute_energy_aggregate(hourly_df, "ME", add_month_year=True)

    def _compute_energy_yearly_view(self, hourly_df):
        return self._compute_energy_aggregate(hourly_df, "YE", add_month_year=False)

    def _build_energy_result_table_tab(self, parent):
        """Build Result table tab: 4 cards (Hourly, Daily, Monthly, Yearly)."""
        container = customtkinter.CTkFrame(parent, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=30, pady=30)
        customtkinter.CTkLabel(container, text="Energy Results Breakdown", font=("Segoe UI", 18, "bold"), text_color=TEXT_PRIMARY).pack(anchor="w", pady=(0, 10))
        customtkinter.CTkLabel(container, text="Click any card to view data table", font=("Segoe UI", 12), text_color=TEXT_SECONDARY).pack(anchor="w", pady=(0, 30))

        hourly_df = getattr(self, "_energy_result_df", None)
        if hourly_df is None or (hasattr(hourly_df, "empty") and hourly_df.empty):
            customtkinter.CTkLabel(container, text="No results. Run Calculate first.", font=("Segoe UI", 12), text_color=TEXT_MUTED).pack(expand=True, pady=24)
            return

        try:
            daily_df = self._compute_energy_daily_view(hourly_df)
        except Exception:
            daily_df = pd.DataFrame()
        try:
            monthly_df = self._compute_energy_monthly_view(hourly_df)
        except Exception:
            monthly_df = pd.DataFrame()
        try:
            yearly_df = self._compute_energy_yearly_view(hourly_df)
        except Exception:
            yearly_df = pd.DataFrame()

        if self._is_tmy_import():
            daily_df = self._drop_empty_aggregate_rows(daily_df)
            monthly_df = self._drop_empty_aggregate_rows(monthly_df)
            yearly_df = self._drop_empty_aggregate_rows(yearly_df)

        # Restrict to core energy columns (no extra import columns)
        def _reorder_cols(df, preferred):
            if df is None or df.empty:
                return df
            ordered = [c for c in preferred if c in df.columns]
            return df[ordered]

        HOURLY_ORDER = ("time", "Ta", "TC(f)", "Psource(f)", "Psource(x)", "Ppv(f)", "Kt", "Ppv(x)")
        DAILY_ORDER = ("time", "Ta", "TC(f)", "Esource(f)", "Esource(x)", "Epv(f)", "Kt", "Epv(x)", "CF(f)", "CF(x)")
        MONTHLY_ORDER = ("time", "Month", "Year", "Ta", "TC(f)", "Esource(f)", "Esource(x)", "Epv(f)", "Kt", "Epv(x)", "CF(f)", "CF(x)")
        YEARLY_ORDER = ("time", "Ta", "TC(f)", "Esource(f)", "Esource(x)", "Epv(f)", "Kt", "Epv(x)", "CF(f)", "CF(x)")

        hourly_df = _reorder_cols(hourly_df, HOURLY_ORDER)
        daily_df = _reorder_cols(daily_df, DAILY_ORDER)
        monthly_df = _reorder_cols(monthly_df, MONTHLY_ORDER)
        yearly_df = _reorder_cols(yearly_df, YEARLY_ORDER)

        cards_frame = customtkinter.CTkFrame(container, fg_color="transparent")
        cards_frame.pack(fill="x")
        cards_frame.grid_columnconfigure((0, 1, 2, 3), weight=1, uniform="cards")
        h_cols = len(hourly_df.columns) if not hourly_df.empty else 0
        cards_data = [
            {"title": "Hourly", "subtitle": f"{fmt_num(len(hourly_df))} records • {h_cols} columns", "icon": "⏱️", "df": hourly_df},
            {"title": "Daily", "subtitle": f"{fmt_num(len(daily_df))} days aggregated", "icon": "📆", "df": daily_df},
            {"title": "Monthly", "subtitle": f"{fmt_num(len(monthly_df))} months aggregated", "icon": "📅", "df": monthly_df},
            {"title": "Yearly", "subtitle": f"{fmt_num(len(yearly_df))} years aggregated", "icon": "📊", "df": yearly_df},
        ]
        for idx, card_data in enumerate(cards_data):
            create_visualize_card(
                cards_frame, card_data["title"], card_data["subtitle"], card_data["icon"], idx,
                lambda d=card_data["df"], t=card_data["title"]: self._open_data_window(t, d, "energy"),
            )

    def _build_energy_graphs_tab(self, parent):
        """Build Graphs tab: modern layout with grouped controls and preview area."""
        main_container = customtkinter.CTkFrame(parent, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=SPACE_LG, pady=SPACE_LG)

        hourly_df = getattr(self, "_energy_result_df", None)
        if hourly_df is None or (hasattr(hourly_df, "empty") and hourly_df.empty):
            customtkinter.CTkLabel(
                main_container, text="No results. Run Calculate first.",
                font=(FONT_FAMILY_TEXT, FONT_SIZE_BODY), text_color=TEXT_MUTED
            ).pack(expand=True, pady=SPACE_XL)
            return

        time_col = next((c for c in hourly_df.columns if "time" in str(c).lower()), None)
        if not time_col:
            customtkinter.CTkLabel(
                main_container, text="No time column in results.",
                font=(FONT_FAMILY_TEXT, FONT_SIZE_BODY), text_color=TEXT_MUTED
            ).pack(expand=True, pady=SPACE_XL)
            return

        t_parsed = self._parse_time_column(hourly_df[time_col])
        years_avail = sorted(t_parsed.dt.year.dropna().unique().astype(int).tolist())
        months_avail = [("January", 1), ("February", 2), ("March", 3), ("April", 4), ("May", 5), ("June", 6),
                       ("July", 7), ("August", 8), ("September", 9), ("October", 10), ("November", 11), ("December", 12)]

        def _safe_int(val, default=None):
            if not val or val == "—":
                return default
            try:
                return int(val)
            except (ValueError, TypeError):
                return default

        def _safe_month(name, default=None):
            if not name or name == "—":
                return default
            for n, num in months_avail:
                if n == name:
                    return num
            return default

        RESOL_OPTS = ["— Select —", "Yearly", "Monthly", "Daily"]
        TIME_YEARLY = ["Hourly", "Daily", "Monthly"]
        TIME_MONTHLY = ["Hourly", "Daily"]
        TIME_DAILY = ["Hourly"]

        # ── Page header ──
        header_row = customtkinter.CTkFrame(main_container, fg_color="transparent")
        header_row.pack(fill="x", pady=(0, 12))
        customtkinter.CTkLabel(
            header_row, text="📈", font=(FONT_FAMILY_DISPLAY, 22),
        ).pack(side="left", padx=(0, 8))
        hdr_text = customtkinter.CTkFrame(header_row, fg_color="transparent")
        hdr_text.pack(side="left")
        customtkinter.CTkLabel(
            hdr_text, text="Energy Charts",
            font=(FONT_FAMILY_DISPLAY, FONT_SIZE_H3, "bold"), text_color=TEXT_PRIMARY,
        ).pack(anchor="w")
        customtkinter.CTkLabel(
            hdr_text, text="Configure period, variable, and generate interactive charts",
            font=(FONT_FAMILY_TEXT, FONT_SIZE_SMALL), text_color=TEXT_SECONDARY,
        ).pack(anchor="w")

        # ── Chart options card ──
        controls = customtkinter.CTkFrame(
            main_container, fg_color=THEME.bg.card, corner_radius=RADIUS_MD,
            border_width=1, border_color=THEME.border.light,
        )
        controls.pack(fill="x", pady=(0, SPACE_MD))

        # Card header
        ctrl_header = customtkinter.CTkFrame(controls, fg_color="transparent")
        ctrl_header.pack(fill="x", padx=20, pady=(16, 4))
        customtkinter.CTkLabel(
            ctrl_header, text="Chart Options",
            font=(FONT_FAMILY_TEXT, 13, "bold"), text_color=TEXT_PRIMARY,
        ).pack(side="left")

        # Divider
        customtkinter.CTkFrame(controls, fg_color=THEME.border.light, height=1).pack(fill="x", padx=20, pady=(8, 0))

        combo_style = dict(
            corner_radius=RADIUS_SM, border_width=1, font=(FONT_FAMILY_TEXT, FONT_SIZE_SMALL),
            dropdown_fg_color=THEME.bg.card, dropdown_hover_color=THEME.bg.selected,
            dropdown_text_color=TEXT_PRIMARY, height=INPUT_HEIGHT,
        )

        # ── Row 1: Period (Resolution + Year/Month/Day dynamic) ──
        period_section = customtkinter.CTkFrame(controls, fg_color="transparent")
        period_section.pack(fill="x", padx=20, pady=(12, 0))

        customtkinter.CTkLabel(
            period_section, text="Period", font=(FONT_FAMILY_TEXT, 12, "bold"),
            text_color=TEXT_PRIMARY,
        ).pack(anchor="w", pady=(0, 8))

        period_fields = customtkinter.CTkFrame(period_section, fg_color="transparent")
        period_fields.pack(fill="x")

        # Resolution
        res_frame = customtkinter.CTkFrame(period_fields, fg_color="transparent")
        res_frame.pack(side="left", padx=(0, 16))
        customtkinter.CTkLabel(
            res_frame, text="Resolution", font=(FONT_FAMILY_TEXT, 11, "bold"),
            text_color=TEXT_SECONDARY,
        ).pack(anchor="w", pady=(0, 4))
        combo_resolution = customtkinter.CTkComboBox(res_frame, values=RESOL_OPTS, width=130, state="readonly", **combo_style)
        combo_resolution.pack(anchor="w")
        combo_resolution.set("— Select —")

        # Year — initially hidden
        yr_frame = customtkinter.CTkFrame(period_fields, fg_color="transparent")
        customtkinter.CTkLabel(
            yr_frame, text="Year", font=(FONT_FAMILY_TEXT, 11, "bold"),
            text_color=TEXT_SECONDARY,
        ).pack(anchor="w", pady=(0, 4))
        year_vals = [str(y) for y in years_avail] if years_avail else ["—"]
        combo_year = customtkinter.CTkComboBox(yr_frame, values=year_vals, width=90, state="readonly", **combo_style)
        combo_year.pack(anchor="w")
        combo_year.set(year_vals[0] if year_vals else "—")

        # Month — initially hidden
        mo_frame = customtkinter.CTkFrame(period_fields, fg_color="transparent")
        customtkinter.CTkLabel(
            mo_frame, text="Month", font=(FONT_FAMILY_TEXT, 11, "bold"),
            text_color=TEXT_SECONDARY,
        ).pack(anchor="w", pady=(0, 4))
        combo_month = customtkinter.CTkComboBox(mo_frame, values=["—"], width=110, state="disabled", **combo_style)
        combo_month.pack(anchor="w")

        # Day — initially hidden
        day_frame = customtkinter.CTkFrame(period_fields, fg_color="transparent")
        customtkinter.CTkLabel(
            day_frame, text="Day", font=(FONT_FAMILY_TEXT, 11, "bold"),
            text_color=TEXT_SECONDARY,
        ).pack(anchor="w", pady=(0, 4))
        combo_day = customtkinter.CTkComboBox(day_frame, values=["—"], width=80, state="disabled", **combo_style)
        combo_day.pack(anchor="w")

        # ── Row 2: Data (Time + Variable + Plot button) ──
        data_section = customtkinter.CTkFrame(controls, fg_color="transparent")
        data_section.pack(fill="x", padx=20, pady=(14, 16))

        customtkinter.CTkLabel(
            data_section, text="Data", font=(FONT_FAMILY_TEXT, 12, "bold"),
            text_color=TEXT_PRIMARY,
        ).pack(anchor="w", pady=(0, 8))

        data_fields = customtkinter.CTkFrame(data_section, fg_color="transparent")
        data_fields.pack(fill="x")

        # Time resolution
        time_frame = customtkinter.CTkFrame(data_fields, fg_color="transparent")
        time_frame.pack(side="left", padx=(0, 16))
        customtkinter.CTkLabel(
            time_frame, text="Time Step", font=(FONT_FAMILY_TEXT, 11, "bold"),
            text_color=TEXT_SECONDARY,
        ).pack(anchor="w", pady=(0, 4))
        combo_time = customtkinter.CTkComboBox(time_frame, values=["—"], width=120, state="disabled", **combo_style)
        combo_time.pack(anchor="w")

        # Variable
        var_frame = customtkinter.CTkFrame(data_fields, fg_color="transparent")
        var_frame.pack(side="left", padx=(0, 24))
        customtkinter.CTkLabel(
            var_frame, text="Variable", font=(FONT_FAMILY_TEXT, 11, "bold"),
            text_color=TEXT_SECONDARY,
        ).pack(anchor="w", pady=(0, 4))
        plottable_hourly = self._energy_plottable_columns(hourly_df)
        var_opts = ["Plot all"] + plottable_hourly if plottable_hourly else ["— No plottable columns —"]
        combo_variable = customtkinter.CTkComboBox(var_frame, values=var_opts, width=160, state="readonly", **combo_style)
        combo_variable.pack(anchor="w")
        combo_variable.set(var_opts[0] if var_opts else "— No plottable columns —")

        # Plot button — prominent at the end
        plot_btn = customtkinter.CTkButton(
            data_fields, text="🔬  Plot Graph",
            width=160, height=BUTTON_HEIGHT_MD, corner_radius=RADIUS_SM,
            font=(FONT_FAMILY_TEXT, FONT_SIZE_BODY, "bold"),
            state="disabled", fg_color=THEME.bg.gray_pale, text_color=TEXT_MUTED,
        )
        plot_btn.pack(side="left", pady=(18, 0))

        def _update_period():
            res = combo_resolution.get()
            if res == "Monthly" or res == "Daily":
                month_values = [m[0] for m in months_avail]
                combo_month.configure(state="readonly", values=month_values)
                cur_mo = combo_month.get()
                if cur_mo not in month_values:
                    combo_month.set(month_values[0])
            else:
                combo_month.configure(state="disabled", values=["—"])
                combo_month.set("—")
            if res == "Daily":
                yr = _safe_int(combo_year.get())
                mn = _safe_month(combo_month.get())
                if yr and mn:
                    try:
                        n = calendar.monthrange(yr, mn)[1]
                        day_values = [str(d) for d in range(1, n + 1)]
                        combo_day.configure(state="readonly", values=day_values)
                        cur_day = combo_day.get()
                        if cur_day not in day_values:
                            combo_day.set("1")
                    except (ValueError, TypeError):
                        combo_day.configure(values=["—"])
                else:
                    combo_day.configure(state="disabled", values=["—"])
            else:
                combo_day.configure(state="disabled", values=["—"])

        def _refresh_vars():
            res = combo_resolution.get()
            t_opt = combo_time.get()
            if not res or res == "— Select —" or not t_opt or t_opt == "—":
                return
            yr = _safe_int(combo_year.get())
            mn = _safe_month(combo_month.get())
            dy = _safe_int(combo_day.get()) if combo_day.cget("state") != "disabled" else None
            plot_df, _, _, _ = self._prepare_energy_plot_data(hourly_df, res, t_opt, yr, mn, dy, time_col)
            opts = self._energy_plottable_columns(plot_df) if plot_df is not None and not plot_df.empty else []
            opts = ["Plot all"] + opts if opts else ["— No data —"]
            combo_variable.configure(values=opts)
            combo_variable.set(opts[0] if opts else "— No data —")

        def _show_hide_period_fields(choice):
            """Dynamically show/hide Year, Month, Day based on resolution."""
            yr_frame.pack_forget()
            mo_frame.pack_forget()
            day_frame.pack_forget()
            if choice in ("Yearly", "Monthly", "Daily"):
                yr_frame.pack(side="left", padx=(0, 16), after=res_frame)
            if choice in ("Monthly", "Daily"):
                mo_frame.pack(side="left", padx=(0, 16), after=yr_frame)
            if choice == "Daily":
                day_frame.pack(side="left", padx=(0, 16), after=mo_frame)

        def on_res(choice):
            combo_time.set("—")
            _show_hide_period_fields(choice)
            if choice and choice != "— Select —":
                combo_time.configure(state="readonly")
                if choice == "Yearly":
                    combo_time.configure(values=TIME_YEARLY)
                    combo_time.set(TIME_YEARLY[0])
                elif choice == "Monthly":
                    combo_time.configure(values=TIME_MONTHLY)
                    combo_time.set(TIME_MONTHLY[0])
                elif choice == "Daily":
                    combo_time.configure(values=TIME_DAILY)
                    combo_time.set(TIME_DAILY[0])
                _update_period()
                _refresh_vars()
            else:
                combo_time.configure(state="disabled", values=["—"])
                combo_month.configure(state="disabled", values=["—"])
                combo_day.configure(state="disabled", values=["—"])
                combo_variable.configure(values=["— Select resolution first —"])
                combo_variable.set("— Select resolution first —")
            _update_btn()

        def on_change(_=None):
            _update_period()
            _refresh_vars()
            _update_btn()

        def _update_btn():
            res = combo_resolution.get()
            var = combo_variable.get()
            t_opt = combo_time.get()
            ok = res and res != "— Select —" and t_opt and t_opt != "—" and var and var not in ("—", "— No plottable columns —", "— No data —", "— Select resolution first —")
            if ok and res in ("Monthly", "Daily"):
                mn = combo_month.get()
                if not mn or mn == "—":
                    ok = False
            if ok and res == "Daily":
                dy = combo_day.get()
                if not dy or dy == "—":
                    ok = False
            if ok:
                plot_btn.configure(state="normal", fg_color=PRIMARY_BLUE, hover_color=PRIMARY_HOVER, text_color="#FFFFFF")
            else:
                plot_btn.configure(state="disabled", fg_color=THEME.bg.gray_pale, text_color=TEXT_MUTED)

        combo_resolution.configure(command=on_res)
        combo_year.configure(command=on_change)
        combo_month.configure(command=on_change)
        combo_day.configure(command=on_change)
        combo_time.configure(command=on_change)
        combo_variable.configure(command=lambda _: _update_btn())

        def do_plot():
            res = combo_resolution.get()
            t_opt = combo_time.get()
            var_display = combo_variable.get()
            if not res or res == "— Select —" or not t_opt or t_opt == "—":
                return
            if not var_display or var_display in ("—", "— No plottable columns —", "— No data —"):
                return
            yr = _safe_int(combo_year.get())
            mn = _safe_month(combo_month.get())
            dy = _safe_int(combo_day.get()) if combo_day.cget("state") != "disabled" else None
            plot_df, x_axis_type, view, title_suffix = self._prepare_energy_plot_data(hourly_df, res, t_opt, yr, mn, dy, time_col)
            if plot_df is None or plot_df.empty:
                show_toast(self, "No data for selected period.", type="warning")
                return
            cols = self._energy_plottable_columns(plot_df)
            if var_display == "Plot all":
                cols_to_plot = cols
            else:
                cols_to_plot = [var_display] if var_display in plot_df.columns else cols
            if not cols_to_plot:
                show_toast(self, "No plottable columns found.", type="warning")
                return
            title = f"{', '.join(cols_to_plot)} vs time" if len(cols_to_plot) > 1 else f"{cols_to_plot[0]} vs time"
            if title_suffix:
                title = f"{title} ({title_suffix})"
            self._show_energy_plot(plot_df, cols_to_plot, title, self, x_axis_type=x_axis_type, view=view)

        plot_btn.configure(command=do_plot)

    def _energy_plottable_columns(self, df):
        """Return core energy output columns only (no import columns). CF only in daily/monthly/yearly."""
        if df is None or df.empty:
            return []
        CORE_ENERGY_COLS = (
            "Ta", "TC(f)", "Psource(f)", "Psource(x)", "Ppv(f)", "Kt", "Ppv(x)",
            "Esource(f)", "Esource(x)", "Epv(f)", "Epv(x)", "CF(f)", "CF(x)",
        )
        return [c for c in CORE_ENERGY_COLS if c in df.columns]

    def _prepare_energy_plot_data(self, hourly_df, resolution, time_opt, year, month, day, time_col):
        """Filter and optionally aggregate hourly data. Returns (df, x_axis_type, view, title_suffix)."""
        t = self._parse_time_column(hourly_df[time_col])
        mask = t.notna()
        if resolution == "Yearly" and year is not None:
            mask = mask & (t.dt.year == year)
        elif resolution == "Monthly" and year is not None and month is not None:
            mask = mask & (t.dt.year == year) & (t.dt.month == month)
        elif resolution == "Daily" and year is not None and month is not None and day is not None:
            mask = mask & (t.dt.year == year) & (t.dt.month == month) & (t.dt.day == day)
        filtered = hourly_df.loc[mask].copy()
        if filtered.empty:
            return (None, "elapsed_hours", "hourly", "")

        if time_opt == "Hourly":
            # Raw hourly: power in W
            x_axis_type = "elapsed_hours"
            view = "hourly"
            if resolution == "Yearly":
                title_suffix = f"Year {year}"
            elif resolution == "Monthly":
                mn_name = calendar.month_abbr[month]
                title_suffix = f"{mn_name} {year}"
            else:
                mn_name = calendar.month_abbr[month]
                title_suffix = f"{mn_name} {day}, {year}"
            return (filtered, x_axis_type, view, title_suffix)

        if time_opt == "Daily":
            # Aggregate to daily: energy in kWh, add CF
            if aggregate_energy_hourly is None:
                return (None, "days", "daily", "")
            agg = aggregate_energy_hourly(filtered, "D", add_month_year=False)
            if agg.empty:
                return (None, "days", "daily", "")
            agg = self._add_cf_columns(agg, "D")
            if self._is_tmy_import():
                agg = self._drop_empty_aggregate_rows(agg)
            x_axis_type = "days"
            view = "daily"
            if resolution == "Yearly":
                title_suffix = f"Year {year}"
            else:
                mn_name = calendar.month_abbr[month]
                title_suffix = f"{mn_name} {year}"
            return (agg, x_axis_type, view, title_suffix)

        if time_opt == "Monthly" and resolution == "Yearly":
            # Aggregate to monthly: energy in kWh, add CF
            if aggregate_energy_hourly is None:
                return (None, "months", "monthly", "")
            agg = aggregate_energy_hourly(filtered, "ME", add_month_year=True)
            if agg.empty:
                return (None, "months", "monthly", "")
            agg = self._add_cf_columns(agg, "ME")
            if self._is_tmy_import():
                agg = self._drop_empty_aggregate_rows(agg)
            x_axis_type = "months"
            view = "monthly"
            title_suffix = f"Year {year}"
            return (agg, x_axis_type, view, title_suffix)

        return (None, "elapsed_hours", "hourly", "")

    def _parse_time_column(self, series):
        """Parse time column like aggregate_energy_hourly: support YYYYMMDD:HHMM, ms, datetime."""
        if pd.api.types.is_datetime64_any_dtype(series):
            return pd.to_datetime(series, utc=True, errors="coerce")
        sample = series.dropna()
        if len(sample) == 0:
            return pd.to_datetime(series, utc=True, errors="coerce")
        first = str(sample.iloc[0])
        if ":" in first and len(first) <= 16:
            return pd.to_datetime(series, format="%Y%m%d:%H%M", utc=True, errors="coerce")
        if pd.api.types.is_numeric_dtype(series):
            return pd.to_datetime(series, unit="ms", utc=True, errors="coerce")
        return pd.to_datetime(series, utc=True, errors="coerce")

    def _show_energy_plot(self, df, col_or_cols, title, parent_frame, x_axis_type="elapsed_hours", view="hourly"):
        """Embed matplotlib plot. x_axis_type: elapsed_hours|days|months|date. view: hourly|daily|monthly for y-axis units."""
        time_col = None
        for c in df.columns:
            if "time" in str(c).lower():
                time_col = c
                break
        if time_col is None and "Month" in df.columns and "Year" in df.columns:
            time_col = next((c for c in df.columns if "time" in str(c).lower()), None)
        if time_col is None:
            show_toast(self, "No time column in data.", type="warning")
            return
        t = self._parse_time_column(df[time_col])
        cols_to_plot = [col_or_cols] if isinstance(col_or_cols, str) else list(col_or_cols)
        cols_to_plot = [c for c in cols_to_plot if c in df.columns]

        if not cols_to_plot:
            show_toast(self, "No plottable columns found.", type="warning")
            return

        try:
            self._show_energy_plot_impl(df, cols_to_plot, t, title, parent_frame, x_axis_type=x_axis_type, view=view)
        except Exception as e:
            show_toast(self, f"Failed to create plot:\n{str(e, type="error")}")
            import traceback
            traceback.print_exc()

    def _show_energy_plot_impl(self, df, cols_to_plot, t, title, parent_frame, x_axis_type="elapsed_hours", view="hourly"):
        """Build figure and open in interactive Toplevel with MATLAB/Simulink-grade tools.

        Features:
        - Zoom (box + scroll wheel), Pan, Home (reset view)
        - Crosshair cursor tracking mouse position
        - Data cursor: click a point to annotate x,y values
        - Grid toggle, Legend toggle
        - Dense tick marks (30 steps)
        - Save figure (PNG/SVG/PDF)
        - Status bar with live cursor coordinates
        """
        try:
            import matplotlib
            matplotlib.use("TkAgg")
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
            from matplotlib.ticker import MaxNLocator, AutoMinorLocator
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
            from matplotlib.widgets import Cursor
        except ImportError:
            show_toast(self, "matplotlib not available.", type="error")
            return

        try:
            plt.style.use("seaborn-v0_8-whitegrid")
        except Exception:
            try:
                plt.style.use("ggplot")
            except Exception:
                pass

        try:
            from .units import get_column_unit
        except ImportError:
            get_column_unit = lambda c, ctx, view=None: ""

        def _get_unit(col):
            return get_column_unit(col, "energy", view=view) if callable(get_column_unit) else ""

        # ── Build x values ──
        t_series = pd.Series(t) if not isinstance(t, pd.Series) else t
        if x_axis_type == "elapsed_hours":
            t_min = t_series.min()
            x_vals = (t_series - t_min).dt.total_seconds() / 3600.0
            x_vals = x_vals.values if hasattr(x_vals, 'values') else np.array(x_vals)
            x_label = "Elapsed time (h)"
            use_datetime_x = False
        elif x_axis_type == "days":
            x_vals = np.arange(1, len(df) + 1)
            x_label = "Day"
            use_datetime_x = False
        elif x_axis_type == "months":
            x_vals = df["Month"].values if "Month" in df.columns else np.arange(1, len(df) + 1)
            x_label = "Month"
            use_datetime_x = False
        else:
            x_vals = t_series
            x_label = "Date"
            use_datetime_x = True

        # ── Color palette ──
        colors = [
            "#007AFF", "#FF3B30", "#34C759", "#FF9500", "#AF52DE",
            "#5AC8FA", "#FF2D55", "#64D2FF", "#FFD60A", "#30D158",
        ]

        ncols = len(cols_to_plot)
        if ncols == 1:
            col = cols_to_plot[0]
            y = pd.to_numeric(df[col], errors="coerce") if col in df.columns else pd.Series(dtype=float)
            if col == "Ta" and (y.isna().all() or col not in df.columns):
                cfg = getattr(self, "import_config", None) or {}
                api_result = cfg.get("api_result") if isinstance(cfg, dict) else None
                import_hourly = api_result.get("hourly_data") if api_result and isinstance(api_result, dict) else None
                if import_hourly is not None and not (hasattr(import_hourly, "empty") and import_hourly.empty):
                    tamb_col = next((tc for tc in ("T2m", "Tamb", "temp", "temperature") if tc in import_hourly.columns), None)
                    if tamb_col is None:
                        for c in import_hourly.columns:
                            if "temp" in str(c).lower() or "t2m" in str(c).lower():
                                tamb_col = c
                                break
                    if tamb_col:
                        import_t_col = next((c for c in import_hourly.columns if "time" in str(c).lower()), None)
                        if import_t_col:
                            imp_t = self._parse_time_column(import_hourly[import_t_col])
                            imp_ta = pd.to_numeric(import_hourly[tamb_col], errors="coerce")
                            right = pd.DataFrame({"_t": imp_t.astype(str), "_ta": imp_ta}).drop_duplicates(subset=["_t"], keep="first")
                            left = pd.DataFrame({"_t": t_series.astype(str)})
                            merged = left.merge(right, on="_t", how="left")
                            y = merged["_ta"].values
            y = pd.Series(y) if not isinstance(y, pd.Series) else y
            valid = t_series.notna() & y.notna()
            if not valid.any():
                show_toast(self, f"No valid data for {col}. Try re-running Calculate to refresh results.", type="warning")
                return
            if use_datetime_x:
                x_ok = t_series[valid].values
            else:
                x_ok = np.asarray(x_vals)[np.asarray(valid, dtype=bool)]
            y_ok = y[valid].values
            fig, ax = plt.subplots(figsize=(14, 7), dpi=110)
            line_color = colors[0]
            ax.plot(x_ok, y_ok, color=line_color, linewidth=1.8, marker=".", markersize=3, markevery=max(1, len(x_ok) // 200))
            unit = _get_unit(col)
            ax.set_ylabel(f"{col} ({unit})" if unit else col, fontsize=12, fontweight="bold")
        else:
            nrows = (ncols + 2) // 3
            nax = min(ncols, 3)
            fig, axes = plt.subplots(nrows, nax, figsize=(14, 4.5 * nrows), dpi=110, squeeze=False)
            axes_flat = axes.flatten()
            for idx, col in enumerate(cols_to_plot):
                ax = axes_flat[idx]
                y = pd.to_numeric(df[col], errors="coerce")
                if col == "Ta" and y.isna().all():
                    cfg = getattr(self, "import_config", None) or {}
                    api_result = cfg.get("api_result") if isinstance(cfg, dict) else None
                    import_hourly = api_result.get("hourly_data") if api_result and isinstance(api_result, dict) else None
                    if import_hourly is not None and not (hasattr(import_hourly, "empty") and import_hourly.empty):
                        tamb_col = next((tc for tc in ("T2m", "Tamb", "temp", "temperature") if tc in import_hourly.columns), None)
                        if tamb_col:
                            import_t_col = next((c for c in import_hourly.columns if "time" in str(c).lower()), None)
                            if import_t_col:
                                imp_t = self._parse_time_column(import_hourly[import_t_col])
                                imp_ta = pd.to_numeric(import_hourly[tamb_col], errors="coerce")
                                right = pd.DataFrame({"_t": imp_t.astype(str), "_ta": imp_ta}).drop_duplicates(subset=["_t"], keep="first")
                                left = pd.DataFrame({"_t": t_series.astype(str)})
                                merged = left.merge(right, on="_t", how="left")
                                y = merged["_ta"].values
                y = pd.Series(y) if not isinstance(y, pd.Series) else y
                valid = t_series.notna() & y.notna()
                line_color = colors[idx % len(colors)]
                if valid.any():
                    if use_datetime_x:
                        xd = t_series[valid].values
                    else:
                        xd = np.asarray(x_vals)[np.asarray(valid, dtype=bool)]
                    yd = y[valid].values
                    ax.plot(xd, yd, color=line_color, linewidth=1.8, marker=".", markersize=3, markevery=max(1, len(xd) // 200))
                unit = _get_unit(col)
                ax.set_ylabel(f"{col} ({unit})" if unit else col, fontsize=10, fontweight="bold")
                ax.grid(True, which="both", linestyle="-", alpha=0.4)
                ax.grid(True, which="minor", linestyle=":", alpha=0.2)
                ax.tick_params(axis="x", rotation=0 if not use_datetime_x else 25)
            for idx in range(ncols, len(axes_flat)):
                axes_flat[idx].set_visible(False)
            for idx in range(1, ncols):
                axes_flat[idx].sharex(axes_flat[0])

        # ── Style all axes ──
        fig.patch.set_facecolor("#FAFBFC")
        axes_list = fig.axes if ncols > 1 else [fig.axes[0]]
        is_multi = ncols > 1
        for ax in axes_list:
            ax.set_facecolor("#FFFFFF")
            ax.set_xlabel(x_label, fontsize=11, fontweight="bold")
            ax.grid(True, which="major", linestyle="-", alpha=0.4, color="#D0D0D0")
            if use_datetime_x:
                ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d %H:%M"))
                ax.xaxis.set_major_locator(mdates.AutoDateLocator())
                ax.tick_params(axis="x", rotation=25, labelsize=9)
            else:
                x_nbins = 8 if is_multi else 10
                ax.xaxis.set_major_locator(MaxNLocator(integer=True, nbins=x_nbins))
                ax.tick_params(axis="x", rotation=0, labelsize=9)
            if not is_multi:
                ax.grid(True, which="minor", linestyle=":", alpha=0.2, color="#E0E0E0")
                ax.minorticks_on()
                ax.xaxis.set_minor_locator(AutoMinorLocator(2))
                ax.yaxis.set_minor_locator(AutoMinorLocator(2))
            ax.yaxis.set_major_locator(MaxNLocator(nbins=10))
            ax.tick_params(axis="y", labelsize=9)
            ax.tick_params(which="minor", length=3, width=0.5)
            ax.tick_params(which="major", length=6, width=1)
            # Subtle spines
            for spine in ax.spines.values():
                spine.set_linewidth(0.5)
                spine.set_color("#C0C0C0")
        fig.suptitle(title, fontsize=14, fontweight="bold", y=0.98)
        fig.tight_layout(rect=[0, 0.03, 1, 0.95])

        # ── Open in dedicated Toplevel with MATLAB-like interactive tools ──
        root = getattr(self, "app", None) or self.winfo_toplevel()
        graph_win = tkinter.Toplevel(root)
        graph_win.title(f"📈 {title[:80]}..." if len(title) > 80 else f"📈 {title}")
        graph_win.geometry("1200x800")
        graph_win.minsize(900, 600)
        graph_win._fig = fig  # Keep figure alive

        # Main frame
        tk_frame = tkinter.Frame(graph_win, bg="#F5F5F7")
        tk_frame.pack(fill="both", expand=True)

        # Canvas
        canvas = FigureCanvasTkAgg(fig, master=tk_frame)
        canvas.draw()
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(side=tkinter.TOP, fill=tkinter.BOTH, expand=True, padx=6, pady=(6, 0))

        # ── Custom tool bar frame with extra buttons ──
        tool_frame = tkinter.Frame(graph_win, bg="#E8E8EC", height=40)
        tool_frame.pack(side=tkinter.BOTTOM, fill=tkinter.X)

        # ── Navigation toolbar (parent must be tool_frame so pack works) ──
        toolbar = NavigationToolbar2Tk(canvas, tool_frame, pack_toolbar=False)
        toolbar.update()
        # Register initial view as "home" so Reset/Home button works
        toolbar.push_current()

        # Navigation toolbar inside our frame
        toolbar.pack(side=tkinter.LEFT, padx=4)

        # Separator
        tkinter.Frame(tool_frame, width=1, bg="#C0C0C0").pack(side=tkinter.LEFT, fill=tkinter.Y, padx=4, pady=4)

        # Grid toggle
        grid_state = [True]
        def toggle_grid():
            grid_state[0] = not grid_state[0]
            for ax in fig.axes:
                ax.grid(grid_state[0], which="major")
                ax.grid(grid_state[0], which="minor")
            grid_btn.config(relief=tkinter.SUNKEN if grid_state[0] else tkinter.RAISED)
            canvas.draw_idle()

        grid_btn = tkinter.Button(tool_frame, text="⊞ Grid", command=toggle_grid,
                                   font=("Segoe UI", 9), relief=tkinter.SUNKEN, bg="#D8D8DC",
                                   padx=6, pady=2)
        grid_btn.pack(side=tkinter.LEFT, padx=2, pady=4)

        # Legend toggle
        legend_state = [False]
        def toggle_legend():
            legend_state[0] = not legend_state[0]
            for ax in fig.axes:
                if legend_state[0]:
                    handles, labels = ax.get_legend_handles_labels()
                    if not labels:
                        labels = [line.get_label() if line.get_label() and not line.get_label().startswith("_") else f"Series {i+1}" for i, line in enumerate(ax.get_lines())]
                        handles = ax.get_lines()
                    if handles:
                        ax.legend(handles, labels, fontsize=9, loc="best", framealpha=0.9)
                else:
                    leg = ax.get_legend()
                    if leg:
                        leg.remove()
            legend_btn.config(relief=tkinter.SUNKEN if legend_state[0] else tkinter.RAISED)
            canvas.draw_idle()

        legend_btn = tkinter.Button(tool_frame, text="◉ Legend", command=toggle_legend,
                                     font=("Segoe UI", 9), relief=tkinter.RAISED, bg="#D8D8DC",
                                     padx=6, pady=2)
        legend_btn.pack(side=tkinter.LEFT, padx=2, pady=4)

        # Crosshair toggle
        crosshair_state = [True]
        crosshair_lines = []
        def toggle_crosshair():
            crosshair_state[0] = not crosshair_state[0]
            for line in crosshair_lines:
                try:
                    line.set_visible(crosshair_state[0])
                except Exception:
                    pass
            cross_btn.config(relief=tkinter.SUNKEN if crosshair_state[0] else tkinter.RAISED)
            canvas.draw_idle()

        # Create crosshair lines on each axis
        for ax in fig.axes:
            hline = ax.axhline(y=0, color="#888888", linewidth=0.7, linestyle="--", alpha=0.6, visible=False)
            vline = ax.axvline(x=0, color="#888888", linewidth=0.7, linestyle="--", alpha=0.6, visible=False)
            crosshair_lines.extend([hline, vline])

        cross_btn = tkinter.Button(tool_frame, text="✚ Crosshair", command=toggle_crosshair,
                                    font=("Segoe UI", 9), relief=tkinter.SUNKEN, bg="#D8D8DC",
                                    padx=6, pady=2)
        cross_btn.pack(side=tkinter.LEFT, padx=2, pady=4)

        # Separator
        tkinter.Frame(tool_frame, width=1, bg="#C0C0C0").pack(side=tkinter.LEFT, fill=tkinter.Y, padx=4, pady=4)

        # Coordinate display
        coord_var = tkinter.StringVar(value="  X: —  |  Y: —  ")
        coord_label = tkinter.Label(tool_frame, textvariable=coord_var,
                                     font=("Consolas", 10), bg="#E8E8EC", fg="#333333",
                                     relief=tkinter.FLAT, padx=10)
        coord_label.pack(side=tkinter.RIGHT, padx=8, pady=4)

        # ── Data cursor annotations ──
        annotations = []

        def on_click(event):
            """Data cursor: click near a data point to annotate its x,y value."""
            if event.inaxes is None:
                return
            if toolbar.mode != "":  # Don't annotate while zooming/panning
                return
            ax = event.inaxes
            for line in ax.get_lines():
                if line in crosshair_lines:
                    continue
                xdata = line.get_xdata()
                ydata = line.get_ydata()
                if len(xdata) == 0:
                    continue
                try:
                    # Find nearest point
                    xdata_f = np.asarray(xdata, dtype=float) if not use_datetime_x else mdates.date2num(xdata)
                    x_click = float(event.xdata) if not use_datetime_x else event.xdata
                    distances = np.abs(xdata_f - x_click)
                    idx = int(np.argmin(distances))
                    x_pt = xdata[idx]
                    y_pt = float(ydata[idx])
                    if use_datetime_x:
                        x_str = mdates.num2date(mdates.date2num(x_pt)).strftime("%Y-%m-%d %H:%M")
                    else:
                        x_str = f"{float(x_pt):.2f}"
                    ann = ax.annotate(
                        f"({x_str}, {y_pt:.3f})",
                        xy=(x_pt, y_pt),
                        xytext=(15, 15),
                        textcoords="offset points",
                        fontsize=8,
                        fontweight="bold",
                        color="#222222",
                        bbox=dict(boxstyle="round,pad=0.3", facecolor="#FFFFCC", edgecolor="#888888", alpha=0.95),
                        arrowprops=dict(arrowstyle="->", color="#888888", lw=1),
                    )
                    annotations.append(ann)
                    # Also mark the point
                    marker = ax.plot(x_pt, y_pt, "o", color="#FF3B30", markersize=6, zorder=10)
                    annotations.extend(marker)
                    canvas.draw_idle()
                    break
                except Exception:
                    pass

        def on_right_click(event):
            """Right-click to clear all data cursor annotations."""
            for ann in annotations:
                try:
                    ann.remove()
                except Exception:
                    pass
            annotations.clear()
            canvas.draw_idle()

        # ── Mouse motion for crosshair + coordinate readout ──
        def on_motion(event):
            if event.inaxes is None:
                coord_var.set("  X: —  |  Y: —  ")
                for line in crosshair_lines:
                    line.set_visible(False)
                canvas.draw_idle()
                return
            x, y = event.xdata, event.ydata
            if x is None or y is None:
                return
            # Update coordinate display
            if use_datetime_x:
                try:
                    x_str = mdates.num2date(x).strftime("%Y-%m-%d %H:%M")
                except Exception:
                    x_str = f"{x:.2f}"
            else:
                x_str = f"{x:.2f}"
            coord_var.set(f"  X: {x_str}  |  Y: {y:.4f}  ")
            # Update crosshair lines
            if crosshair_state[0]:
                for i, ax in enumerate(fig.axes):
                    idx_base = i * 2
                    if idx_base + 1 < len(crosshair_lines):
                        crosshair_lines[idx_base].set_ydata([y])
                        crosshair_lines[idx_base].set_visible(ax == event.inaxes)
                        crosshair_lines[idx_base + 1].set_xdata([x])
                        crosshair_lines[idx_base + 1].set_visible(ax == event.inaxes)
                canvas.draw_idle()

        # ── Scroll wheel zoom ──
        def on_scroll(event):
            if event.inaxes is None:
                return
            ax = event.inaxes
            scale = 1.15
            if event.button == "up":
                factor = 1.0 / scale
            elif event.button == "down":
                factor = scale
            else:
                return
            xmin, xmax = ax.get_xlim()
            ymin, ymax = ax.get_ylim()
            x_center = event.xdata
            y_center = event.ydata
            if x_center is not None:
                new_xrange = (xmax - xmin) * factor
                ax.set_xlim(x_center - new_xrange / 2, x_center + new_xrange / 2)
            if y_center is not None:
                new_yrange = (ymax - ymin) * factor
                ax.set_ylim(y_center - new_yrange / 2, y_center + new_yrange / 2)
            canvas.draw_idle()

        # ── Connect events ──
        canvas.mpl_connect("motion_notify_event", on_motion)
        canvas.mpl_connect("button_press_event", on_click)
        canvas.mpl_connect("button_press_event", lambda e: on_right_click(e) if e.button == 3 else None)
        canvas.mpl_connect("scroll_event", on_scroll)

        graph_win.lift()
        graph_win.focus_force()

    def _get_energy_export_dfs(self):
        """Return dict of reordered hourly/daily/monthly/yearly DataFrames for export."""
        hourly_df = getattr(self, "_energy_result_df", None)
        if hourly_df is None or (hasattr(hourly_df, "empty") and hourly_df.empty):
            return {}

        def _reorder_cols(df, preferred):
            if df is None or (hasattr(df, "empty") and df.empty):
                return df
            ordered = [c for c in preferred if c in df.columns]
            return df[ordered] if ordered else df

        HOURLY_ORDER = ("time", "Ta", "TC(f)", "Psource(f)", "Psource(x)", "Ppv(f)", "Kt", "Ppv(x)")
        DAILY_ORDER = ("time", "Ta", "TC(f)", "Esource(f)", "Esource(x)", "Epv(f)", "Kt", "Epv(x)", "CF(f)", "CF(x)")
        MONTHLY_ORDER = ("time", "Month", "Year", "Ta", "TC(f)", "Esource(f)", "Esource(x)", "Epv(f)", "Kt", "Epv(x)", "CF(f)", "CF(x)")
        YEARLY_ORDER = ("time", "Ta", "TC(f)", "Esource(f)", "Esource(x)", "Epv(f)", "Kt", "Epv(x)", "CF(f)", "CF(x)")
        try:
            daily_df = self._compute_energy_daily_view(hourly_df)
        except Exception:
            daily_df = pd.DataFrame()
        try:
            monthly_df = self._compute_energy_monthly_view(hourly_df)
        except Exception:
            monthly_df = pd.DataFrame()
        try:
            yearly_df = self._compute_energy_yearly_view(hourly_df)
        except Exception:
            yearly_df = pd.DataFrame()
        if self._is_tmy_import():
            daily_df = self._drop_empty_aggregate_rows(daily_df)
            monthly_df = self._drop_empty_aggregate_rows(monthly_df)
            yearly_df = self._drop_empty_aggregate_rows(yearly_df)
        return {
            "hourly": _reorder_cols(hourly_df, HOURLY_ORDER),
            "daily": _reorder_cols(daily_df, DAILY_ORDER),
            "monthly": _reorder_cols(monthly_df, MONTHLY_ORDER),
            "yearly": _reorder_cols(yearly_df, YEARLY_ORDER),
        }

    def _write_energy_df_to_excel_sheet(self, ws, df, get_unit):
        """Write DataFrame to openpyxl sheet with row1=header, row2=units, then data."""
        if df is None or (hasattr(df, "empty") and df.empty):
            ws.append(["No data"])
            return
        units_row = [get_unit(c) for c in df.columns]
        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=1, column=j, value=col)
        for j, unit in enumerate(units_row, start=1):
            ws.cell(row=2, column=j, value=unit)
        for i, row in df.iterrows():
            out = []
            for v in row:
                if hasattr(v, "strftime"):
                    out.append(v.strftime("%Y-%m-%d %H:%M"))
                elif v is None or (isinstance(v, float) and v != v):
                    out.append("")
                else:
                    out.append(v)
            ws.append(out)

    def _write_energy_metadata_to_sheet(self, ws):
        """Write Energy calculation configuration to sheet (key-value rows)."""
        from datetime import datetime
        cfg = getattr(self, "import_config", None) or {}
        peak_kw = self._get_peak_power_kw()
        panel = (self.energy_panel_combo.get() or "N/A") if hasattr(self, "energy_panel_combo") else "N/A"
        tc_id = (self.energy_tc_combo.get() or "N/A") if hasattr(self, "energy_tc_combo") else "N/A"
        ppv_id = (self.energy_ppv_combo.get() or "N/A") if hasattr(self, "energy_ppv_combo") else "N/A"
        if hasattr(self, "energy_mounting_combo"):
            mount = (self.energy_mounting_combo.get() or "N/A")
        else:
            mount = _derive_mounting_for_energy(cfg, cfg.get("source", ""))
            if isinstance(mount, str) and mount != "N/A":
                mount = mount.replace("_", " ").title()
            else:
                mount = cfg.get("mounting_type") or cfg.get("mounting") or "N/A"
        rows = [
            ["Energy calculation – Configuration", ""],
            ["", ""],
            ["Parameter", "Value"],
            ["Peak power (kW)", str(peak_kw)],
            ["Panel", str(panel)],
            ["TC model", str(tc_id)],
            ["Ppv model", str(ppv_id)],
            ["Mounting", str(mount)],
            ["Source", cfg.get("source", "N/A")],
            ["Export date", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ]
        for r in rows:
            ws.append(r)

    def _on_energy_export_excel_tab(self):
        """Export selected Energy data to Excel: one sheet per selected."""
        lbl = getattr(self, "energy_export_tab_status_label", None)
        if not any([
            getattr(self, "energy_export_hourly_var", tkinter.BooleanVar(value=False)).get(),
            getattr(self, "energy_export_daily_var", tkinter.BooleanVar(value=False)).get(),
            getattr(self, "energy_export_monthly_var", tkinter.BooleanVar(value=False)).get(),
            getattr(self, "energy_export_yearly_var", tkinter.BooleanVar(value=False)).get(),
            getattr(self, "energy_export_metadata_var", tkinter.BooleanVar(value=False)).get(),
        ]):
            if lbl and lbl.winfo_exists():
                lbl.configure(text="Select at least one data type to export.", text_color=ERROR_RED)
            show_toast(self, "Select at least one data type to export.", type="info")
            return
        hourly_df = getattr(self, "_energy_result_df", None)
        if hourly_df is None or (hasattr(hourly_df, "empty") and hourly_df.empty):
            if lbl and lbl.winfo_exists():
                lbl.configure(text="No results to export. Run Calculate first.", text_color=ERROR_RED)
            show_toast(self, "No results to export. Run Calculate first.", type="info")
            return
        path = filedialog.asksaveasfilename(
            title="Export Energy results",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            from .units import get_column_unit
        except ImportError:
            get_column_unit = lambda c, ctx: ""
        get_unit = lambda c: get_column_unit(c, "energy")
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            dfs = self._get_energy_export_dfs()
            if self.energy_export_hourly_var.get() and dfs.get("hourly") is not None and not dfs["hourly"].empty:
                ws = wb.create_sheet("Hourly Data")
                self._write_energy_df_to_excel_sheet(ws, dfs["hourly"], get_unit)
            if self.energy_export_daily_var.get() and dfs.get("daily") is not None and not dfs["daily"].empty:
                ws = wb.create_sheet("Daily Summary")
                self._write_energy_df_to_excel_sheet(ws, dfs["daily"], get_unit)
            if self.energy_export_monthly_var.get() and dfs.get("monthly") is not None and not dfs["monthly"].empty:
                ws = wb.create_sheet("Monthly Summary")
                self._write_energy_df_to_excel_sheet(ws, dfs["monthly"], get_unit)
            if self.energy_export_yearly_var.get() and dfs.get("yearly") is not None and not dfs["yearly"].empty:
                ws = wb.create_sheet("Yearly Summary")
                self._write_energy_df_to_excel_sheet(ws, dfs["yearly"], get_unit)
            if self.energy_export_metadata_var.get():
                ws = wb.create_sheet("Metadata")
                self._write_energy_metadata_to_sheet(ws)
            wb.save(path)
            fname = path.split(chr(92))[-1].split("/")[-1]
            if lbl and lbl.winfo_exists():
                lbl.configure(text=f"✓ Exported to {fname}", text_color=SUCCESS_GREEN)
            show_toast(self, f"Exported to {fname}", type="info")
        except Exception as e:
            if lbl and lbl.winfo_exists():
                lbl.configure(text=f"Export failed: {str(e)}", text_color=ERROR_RED)
            show_toast(self, str(e, type="error"))

    def _on_energy_export_csv_tab(self):
        """Export selected Energy data to multiple CSV files in a chosen folder."""
        lbl = getattr(self, "energy_export_tab_status_label", None)
        if not any([
            getattr(self, "energy_export_hourly_var", tkinter.BooleanVar(value=False)).get(),
            getattr(self, "energy_export_daily_var", tkinter.BooleanVar(value=False)).get(),
            getattr(self, "energy_export_monthly_var", tkinter.BooleanVar(value=False)).get(),
            getattr(self, "energy_export_yearly_var", tkinter.BooleanVar(value=False)).get(),
            getattr(self, "energy_export_metadata_var", tkinter.BooleanVar(value=False)).get(),
        ]):
            if lbl and lbl.winfo_exists():
                lbl.configure(text="Select at least one data type to export.", text_color=ERROR_RED)
            show_toast(self, "Select at least one data type to export.", type="info")
            return
        hourly_df = getattr(self, "_energy_result_df", None)
        if hourly_df is None or (hasattr(hourly_df, "empty") and hourly_df.empty):
            if lbl and lbl.winfo_exists():
                lbl.configure(text="No results to export. Run Calculate first.", text_color=ERROR_RED)
            show_toast(self, "No results to export. Run Calculate first.", type="info")
            return
        folder_path = filedialog.askdirectory(title="Select folder to save CSV files")
        if not folder_path:
            return
        try:
            from .units import get_column_unit
        except ImportError:
            get_column_unit = lambda c, ctx: ""
        get_unit = lambda c: get_column_unit(c, "energy")
        from datetime import datetime
        import csv
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"Energy_{timestamp}"
        exported = []
        dfs = self._get_energy_export_dfs()

        def write_csv_with_units(filepath, df):
            if df is None or (hasattr(df, "empty") and df.empty):
                return 0
            units_row = [get_unit(c) for c in df.columns]
            with open(filepath, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                w.writerow(list(df.columns))
                w.writerow(units_row)
                for _, row in df.iterrows():
                    out = []
                    for v in row:
                        if hasattr(v, "strftime"):
                            out.append(v.strftime("%Y-%m-%d %H:%M"))
                        elif v is None or (isinstance(v, float) and v != v):
                            out.append("")
                        else:
                            out.append(v)
                    w.writerow(out)
            return len(df)

        try:
            if self.energy_export_hourly_var.get() and dfs.get("hourly") is not None and not dfs["hourly"].empty:
                p = os.path.join(folder_path, f"{base_name}_Hourly_Data.csv")
                write_csv_with_units(p, dfs["hourly"])
                exported.append("Hourly Data")
            if self.energy_export_daily_var.get() and dfs.get("daily") is not None and not dfs["daily"].empty:
                p = os.path.join(folder_path, f"{base_name}_Daily_Summary.csv")
                write_csv_with_units(p, dfs["daily"])
                exported.append("Daily Summary")
            if self.energy_export_monthly_var.get() and dfs.get("monthly") is not None and not dfs["monthly"].empty:
                p = os.path.join(folder_path, f"{base_name}_Monthly_Summary.csv")
                write_csv_with_units(p, dfs["monthly"])
                exported.append("Monthly Summary")
            if self.energy_export_yearly_var.get() and dfs.get("yearly") is not None and not dfs["yearly"].empty:
                p = os.path.join(folder_path, f"{base_name}_Yearly_Summary.csv")
                write_csv_with_units(p, dfs["yearly"])
                exported.append("Yearly Summary")
            if self.energy_export_metadata_var.get():
                p = os.path.join(folder_path, f"{base_name}_Metadata.csv")
                cfg = getattr(self, "import_config", None) or {}
                peak_kw = self._get_peak_power_kw()
                panel = (self.energy_panel_combo.get() or "N/A") if hasattr(self, "energy_panel_combo") else "N/A"
                with open(p, "w", encoding="utf-8", newline="") as f:
                    w = csv.writer(f)
                    w.writerow(["Parameter", "Value"])
                    w.writerow(["Peak power (kW)", str(peak_kw)])
                    w.writerow(["Panel", str(panel)])
                    w.writerow(["Export date", datetime.now().strftime("%Y-%m-%d %H:%M")])
                exported.append("Metadata")
            if lbl and lbl.winfo_exists():
                lbl.configure(
                    text=f"✓ Saved {len(exported)} file(s): {', '.join(exported)}",
                    text_color=SUCCESS_GREEN,
                )
            show_toast(self, f"Saved {len(exported, type="info")} file(s) to folder.")
        except Exception as e:
            if lbl and lbl.winfo_exists():
                lbl.configure(text=f"CSV export failed: {str(e)}", text_color=ERROR_RED)
            show_toast(self, str(e, type="error"))

    def _build_energy_export_tab(self, parent):
        """Build Export tab: choices (Hourly, Daily, Monthly, Yearly, Metadata) and CSV/Excel export."""
        scroll = customtkinter.CTkScrollableFrame(parent, fg_color="transparent", scrollbar_button_color=THEME.border.gray, scrollbar_button_hover_color=THEME.border.medium)
        scroll.pack(fill="both", expand=True, padx=SPACE_LG, pady=SPACE_LG)

        # ── Header ──
        header_row = customtkinter.CTkFrame(scroll, fg_color="transparent")
        header_row.pack(fill="x", pady=(0, 12))
        customtkinter.CTkLabel(
            header_row, text="📦", font=(FONT_FAMILY_DISPLAY, 22),
        ).pack(side="left", padx=(0, 8))
        hdr_text = customtkinter.CTkFrame(header_row, fg_color="transparent")
        hdr_text.pack(side="left")
        customtkinter.CTkLabel(
            hdr_text, text="Export Energy Results",
            font=(FONT_FAMILY_DISPLAY, FONT_SIZE_H3, "bold"), text_color=TEXT_PRIMARY,
        ).pack(anchor="w")
        customtkinter.CTkLabel(
            hdr_text, text="Select what to export, then save to Excel or CSV. Includes a units row for data interpretation.",
            font=(FONT_FAMILY_TEXT, FONT_SIZE_SMALL), text_color=TEXT_SECONDARY,
        ).pack(anchor="w")

        # ── Select data to export (like Step 1), including Daily ──
        options_frame = customtkinter.CTkFrame(scroll, fg_color="transparent")
        options_frame.pack(fill="x", pady=(SPACE_SM, 0))
        customtkinter.CTkLabel(
            options_frame, text="Select data to export:",
            font=(FONT_FAMILY_TEXT, 13), text_color=TEXT_SECONDARY,
        ).pack(anchor="w", pady=(0, 10))
        self.energy_export_hourly_var = tkinter.BooleanVar(value=True)
        self.energy_export_daily_var = tkinter.BooleanVar(value=True)
        self.energy_export_monthly_var = tkinter.BooleanVar(value=True)
        self.energy_export_yearly_var = tkinter.BooleanVar(value=True)
        self.energy_export_metadata_var = tkinter.BooleanVar(value=True)
        customtkinter.CTkCheckBox(
            options_frame, text="Hourly time series data",
            variable=self.energy_export_hourly_var, font=(FONT_FAMILY_TEXT, 12),
        ).pack(anchor="w", pady=5)
        customtkinter.CTkCheckBox(
            options_frame, text="Daily aggregated summary",
            variable=self.energy_export_daily_var, font=(FONT_FAMILY_TEXT, 12),
        ).pack(anchor="w", pady=5)
        customtkinter.CTkCheckBox(
            options_frame, text="Monthly aggregated summary",
            variable=self.energy_export_monthly_var, font=(FONT_FAMILY_TEXT, 12),
        ).pack(anchor="w", pady=5)
        customtkinter.CTkCheckBox(
            options_frame, text="Yearly totals and KPIs",
            variable=self.energy_export_yearly_var, font=(FONT_FAMILY_TEXT, 12),
        ).pack(anchor="w", pady=5)
        customtkinter.CTkCheckBox(
            options_frame, text="Calculation configuration & metadata",
            variable=self.energy_export_metadata_var, font=(FONT_FAMILY_TEXT, 12),
        ).pack(anchor="w", pady=5)
        self.energy_export_tab_status_label = customtkinter.CTkLabel(
            scroll, text="Ready to export to local PC",
            font=(FONT_FAMILY_TEXT, 12), text_color=TEXT_SECONDARY,
        )
        self.energy_export_tab_status_label.pack(anchor="w", pady=(SPACE_SM, SPACE_MD))

        # ── Export card ──
        card = customtkinter.CTkFrame(
            scroll, fg_color=THEME.bg.card, corner_radius=RADIUS_MD,
            border_width=1, border_color=THEME.border.light,
        )
        card.pack(fill="x", pady=(0, SPACE_MD))

        card_inner = customtkinter.CTkFrame(card, fg_color="transparent")
        card_inner.pack(fill="x", padx=20, pady=20)

        customtkinter.CTkLabel(
            card_inner, text="Choose export format",
            font=(FONT_FAMILY_TEXT, 13, "bold"), text_color=TEXT_PRIMARY,
        ).pack(anchor="w", pady=(0, 4))
        customtkinter.CTkLabel(
            card_inner, text="Excel: one sheet per selected data. CSV: one file per selected (save to a folder).",
            font=(FONT_FAMILY_TEXT, FONT_SIZE_SMALL), text_color=TEXT_MUTED,
        ).pack(anchor="w", pady=(0, 16))

        btn_row = customtkinter.CTkFrame(card_inner, fg_color="transparent")
        btn_row.pack(fill="x")
        customtkinter.CTkButton(
            btn_row, text="📊  Export to Excel (.xlsx)",
            command=self._on_energy_export_excel_tab,
            fg_color=PRIMARY_BLUE, hover_color=PRIMARY_HOVER, text_color="#FFFFFF",
            width=200, height=42, corner_radius=RADIUS_SM, font=(FONT_FAMILY_TEXT, 13, "bold"),
        ).pack(side="left", padx=(0, 12))
        customtkinter.CTkButton(
            btn_row, text="📄  Export to CSV (.csv)",
            command=self._on_energy_export_csv_tab,
            fg_color=THEME.bg.gray_pale, border_width=1, border_color=THEME.border.light,
            text_color=TEXT_PRIMARY,
            width=200, height=42, corner_radius=RADIUS_SM, font=(FONT_FAMILY_TEXT, 12),
        ).pack(side="left")

    def _on_energy_export_dialog(self, ext):
        """Export Energy results to file (CSV or Excel) with units row."""
        df = getattr(self, "_energy_result_df", None)
        if df is None or (hasattr(df, "empty") and df.empty):
            show_toast(self, "No results to export.", type="info")
            return
        path = filedialog.asksaveasfilename(
            title="Export Energy results",
            defaultextension=ext,
            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            from .units import get_column_unit
        except ImportError:
            get_column_unit = lambda c, ctx: ""
        try:
            import pandas as pd
            units_row = [get_column_unit(c, "energy") for c in df.columns]
            if path.lower().endswith(".xlsx"):
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name="Energy", index=False, startrow=2, header=False)
                    ws = writer.sheets["Energy"]
                    for j, col in enumerate(df.columns, start=1):
                        ws.cell(row=1, column=j, value=col)
                    for j, unit in enumerate(units_row, start=1):
                        ws.cell(row=2, column=j, value=unit)
            else:
                import csv
                units_row = [get_column_unit(c, "energy") for c in df.columns]
                with open(path, "w", encoding="utf-8", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerow(list(df.columns))
                    writer.writerow(units_row)
                    for _, row in df.iterrows():
                        out = []
                        for v in row:
                            if hasattr(v, "strftime"):
                                out.append(v.strftime("%Y-%m-%d %H:%M"))
                            elif v is None or (isinstance(v, float) and v != v):
                                out.append("")
                            else:
                                out.append(v)
                        writer.writerow(out)
            show_toast(self, f"Exported to {path.split(chr(92, type="info"))[-1].split('/')[-1]}")
        except Exception as e:
            show_toast(self, str(e, type="error"))

    def _on_energy_export(self):
        """Export Energy results table to CSV or Excel (legacy, kept for compatibility)."""
        df = getattr(self, "_energy_result_df", None)
        if df is None or (hasattr(df, "empty") and df.empty):
            show_toast(self, "No results to export. Run Calculate first.", type="info")
            return
        path = filedialog.asksaveasfilename(
            title="Export Energy results",
            defaultextension=".csv",
            filetypes=[
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx"),
                ("All files", "*.*"),
            ],
        )
        if not path:
            return
        try:
            if path.lower().endswith(".xlsx"):
                df.to_excel(path, index=False)
            else:
                df.to_csv(path, index=False, encoding="utf-8")
            self.energy_status_label.configure(
                text=f"Exported to {path.split('/')[-1].split(chr(92))[-1]}",
                text_color=SUCCESS_GREEN,
            )
        except Exception as e:
            show_toast(self, str(e, type="error"))
            self.energy_status_label.configure(text=f"Export failed: {e}", text_color=ERROR_RED)

# ui/source/step2_visualize.py
# Step 2: Visualize & Export Data - Visualization, KPIs, and Export functionality

import customtkinter
import tkinter
import pandas as pd
import numpy as np
import calendar
import tkinter.ttk as ttk
from tkinter import filedialog
from ..toast import show_toast
from datetime import datetime
from .constants import *
from .units import get_column_unit, format_column_header
from .components import create_kpi_card, render_dataframe_table, create_visualize_card, add_summary_row

# Optional Excel export support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class Step2VisualizeMixin:
    """
    Step 2: Visualize Data functionality
    Contains tabs for Import Summary, Data Tables, and Export
    """
    
    def _build_step2_outputs(self, parent=None):
        """Build the Visualize step UI with Apple-style segmented tabs.
        If parent is given, build inside that container (e.g. popup); else use self.content_inner.
        """
        self._ensure_view_data()
        container = parent if parent is not None else self.content_inner

        for widget in container.winfo_children():
            widget.destroy()

        # Styled tabview – visible selected state
        tabview = customtkinter.CTkTabview(
            container, fg_color="transparent",
            segmented_button_fg_color=THEME.bg.hover,
            segmented_button_selected_color=THEME.primary.blue,
            segmented_button_selected_hover_color=THEME.primary.blue_hover,
            segmented_button_unselected_color=THEME.bg.hover,
            segmented_button_unselected_hover_color=THEME.border.light,
            text_color="#FFFFFF",
            text_color_disabled=THEME.text.secondary,
            corner_radius=10,
        )
        tabview.pack(fill="both", expand=True, padx=8, pady=(8, 4))
        
        tab1 = tabview.add("Import Summary")
        tab2 = tabview.add("Data Tables")
        tab3 = tabview.add("Export")
        
        self._build_import_summary_tab(tab1)
        self._build_data_tables_tab(tab2)
        self._build_graphs_tab(tab3)

    def _build_import_summary_tab(self, parent):
        """Build the Import Summary tab with scrollable content."""
        scroll_container = customtkinter.CTkScrollableFrame(
            parent, 
            fg_color="transparent",
            scrollbar_button_color=THEME.border.gray, 
            scrollbar_button_hover_color=THEME.border.medium
        )
        scroll_container.pack(fill="both", expand=True, padx=30, pady=30)
        
        customtkinter.CTkLabel(
            scroll_container,
            text="Import Configuration",
            font=(FONT_FAMILY_DISPLAY, 18, "bold"),
            text_color=TEXT_PRIMARY
        ).pack(anchor="w", pady=(0, 20))
        
        # Determine data type badge
        config = self.import_config or {}
        
        if self._view_is_tmy:
            data_type_text = "☀️ Irradiation Data (TMY)"
            data_type_color = THEME.status.warning_light
        else:
            data_type_text = "⚡ PV Power Data"
            data_type_color = THEME.status.success_alt
            
        badge = customtkinter.CTkLabel(
            scroll_container,
            text=data_type_text,
            font=(FONT_FAMILY_TEXT, 14, "bold"),
            text_color=data_type_color,
            fg_color=THEME.bg.gray_light,
            corner_radius=8,
            padx=15,
            pady=8
        )
        badge.pack(anchor="w", pady=(0, 20))
        
        details_frame = customtkinter.CTkFrame(scroll_container, fg_color=THEME.bg.gray_lighter, corner_radius=10)
        details_frame.pack(fill="x", pady=10)
        
        def add_config_row(frame, icon, label, value):
            row = customtkinter.CTkFrame(frame, fg_color="transparent")
            row.pack(fill="x", padx=20, pady=8)
            customtkinter.CTkLabel(row, text=icon, font=(FONT_FAMILY_TEXT, 16), width=30).pack(side="left")
            customtkinter.CTkLabel(row, text=label, font=(FONT_FAMILY_TEXT, 12), text_color=TEXT_SECONDARY, width=120).pack(side="left")
            customtkinter.CTkLabel(row, text=str(value), font=(FONT_FAMILY_TEXT, 12, "bold"), text_color=TEXT_PRIMARY).pack(side="left", padx=(10, 0))
        
        add_config_row(details_frame, "📍", "Location:", 
                       f"Lat {config.get('latitude', 'N/A')}°, Lon {config.get('longitude', 'N/A')}°")
        
        # Data source display
        if config.get('source') == "NINJA":
            add_config_row(details_frame, "📊", "Data Source:", f"{config.get('source', 'N/A')} (PV)")
        else:
            add_config_row(details_frame, "📊", "Data Source:", 
                           f"{config.get('source', 'N/A')} ({config.get('pvgis_mode', 'N/A')})")
        
        if config.get('source') == "PVGIS" and config.get('pvgis_mode') == "TMY":
            add_config_row(details_frame, "📅", "Time Period:", f"TMY - {config.get('tmy_database', 'N/A')}")
        elif config.get('source') == "PVGIS":
            add_config_row(details_frame, "📅", "Time Period:", f"{config.get('start_year', 'N/A')} - {config.get('end_year', 'N/A')}")
        else:
            add_config_row(details_frame, "📅", "Time Period:", str(config.get('year', 'N/A')))
        
        # System size
        if config.get('source') == "PVGIS":
            add_config_row(details_frame, "⚡", "System Size:", f"{config.get('peak_power_kwp', 'N/A')} kWp")
        else:
            add_config_row(details_frame, "⚡", "System Size:", f"{config.get('capacity_kw', 'N/A')} kW")
        
        # PV-specific fields (Ninja PV includes wind speed at PV height)
        if config.get('source') == "NINJA":
            add_config_row(details_frame, "📏", "PV height:", f"{config.get('pv_height_m', 'N/A')} m")
        if config.get('use_optimal_angles') or (config.get('optimize_slope') or config.get('optimize_slope_azimuth')):
            tilt_azimuth_text = "Optimal (calculated by API)"
        else:
            tilt_val = config.get('slope', config.get('tilt', 'N/A'))
            azimuth_val = config.get('azimuth', 'N/A')
            tilt_azimuth_text = f"{tilt_val}° / {azimuth_val}°"
        add_config_row(details_frame, "📐", "Tilt/Azimuth:", tilt_azimuth_text)
        add_config_row(details_frame, "🔄", "Tracking:", config.get('mounting_type', config.get('tracking', 'N/A')))
        
        count_frame = customtkinter.CTkFrame(scroll_container, fg_color="transparent")
        count_frame.pack(fill="x", pady=(20, 0))
        customtkinter.CTkLabel(
            count_frame,
            text=f"📈 Total Records: {fmt_num(len(self._view_hourly))} hourly data points",
            font=(FONT_FAMILY_TEXT, 14),
            text_color=THEME.text.gray_dark
        ).pack(anchor="w")
        
        kpi_frame = customtkinter.CTkFrame(scroll_container, fg_color="transparent")
        kpi_frame.pack(fill="x", pady=(30, 20))
        kpi_frame.grid_columnconfigure((0, 1, 2), weight=1)
        
        # Determine KPI labels based on data type
        if self._view_is_tmy:
            annual_label = "Annual Irradiation"
            annual_unit = "kWh/m²"
            yield_label = "Specific Irradiation"
            yield_desc = "kWh / (kWp·m²)"
            cf_desc = "of max possible"
        else:
            annual_label = "Annual Energy"
            annual_unit = "kWh"
            yield_label = "Specific Yield"
            yield_desc = "kWh / kWp"
            cf_desc = "capacity factor"
        
        create_kpi_card(
            kpi_frame, 
            annual_label,
            f"{fmt_num(getattr(self, '_kpi_annual_energy', 0), 0)}",
            annual_unit,
            0, 0
        )
        
        create_kpi_card(
            kpi_frame,
            yield_label,
            fmt_num(self._kpi_specific_yield, 0),
            yield_desc,
            0, 1
        )
        
        create_kpi_card(
            kpi_frame,
            "Capacity Factor",
            f"{self._kpi_capacity_factor:.1f}",
            cf_desc,
            0, 2
        )

    def _build_data_tables_tab(self, parent):
        """Build cards with DYNAMIC record counts from actual dataframes."""
        container = customtkinter.CTkFrame(parent, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=30, pady=30)
        
        customtkinter.CTkLabel(
            container,
            text="Imported Data Breakdown",
            font=(FONT_FAMILY_DISPLAY, 18, "bold"),
            text_color=TEXT_PRIMARY
        ).pack(anchor="w", pady=(0, 10))
        
        customtkinter.CTkLabel(
            container, 
            text="Click any card to view detailed data table",
            font=(FONT_FAMILY_TEXT, 12),
            text_color=TEXT_SECONDARY
        ).pack(anchor="w", pady=(0, 30))
        
        hourly_rows = len(self._view_hourly) if hasattr(self, '_view_hourly') else 0
        monthly_rows = len(self._view_monthly) if hasattr(self, '_view_monthly') else 0
        yearly_rows = len(self._view_yearly) if hasattr(self, '_view_yearly') else 0
        hourly_cols = len(self._view_hourly.columns) if hasattr(self, '_view_hourly') else 0
        
        
        cards_frame = customtkinter.CTkFrame(container, fg_color="transparent")
        cards_frame.pack(fill="x")
        cards_frame.grid_columnconfigure((0, 1, 2), weight=1, uniform="cards")
        
        cards_data = [
            {
                "title": "Hourly Data",
                "subtitle": f"{fmt_num(hourly_rows)} records • {hourly_cols} columns",
                "icon": "⏱️",
                "df": self._view_hourly
            },
            {
                "title": "Monthly Data", 
                "subtitle": f"{monthly_rows} months aggregated",
                "icon": "📅",
                "df": self._view_monthly
            },
            {
                "title": "Yearly Data",
                "subtitle": f"{yearly_rows} years aggregated", 
                "icon": "📊",
                "df": self._view_yearly
            }
        ]
        
        for idx, card_data in enumerate(cards_data):
            title = card_data["title"]
            df = card_data["df"]
            create_visualize_card(
                cards_frame,
                title,
                card_data["subtitle"],
                card_data["icon"],
                idx,
                lambda d=df, t=title: self._open_data_window(t, d, t.lower())
            )

    def _build_graphs_tab(self, parent):
        """Build the Export tab with local PC save functionality."""
        # Use scrollable frame like Import Summary tab
        scroll_container = customtkinter.CTkScrollableFrame(
            parent, 
            fg_color="transparent",
            scrollbar_button_color=THEME.border.gray, 
            scrollbar_button_hover_color=THEME.border.medium
        )
        scroll_container.pack(fill="both", expand=True, padx=30, pady=30)
        
        # Title
        title = customtkinter.CTkLabel(
            scroll_container,
            text="Export Solar Data",
            font=(FONT_FAMILY_DISPLAY, 18, "bold"),
            text_color=TEXT_PRIMARY
        )
        title.pack(anchor="w", pady=(0, 20))
        
        # Options Frame
        options_frame = customtkinter.CTkFrame(scroll_container, fg_color="transparent")
        options_frame.pack(fill="x", pady=(0, 20))
        
        customtkinter.CTkLabel(
            options_frame,
            text="Select data to export:",
            font=(FONT_FAMILY_TEXT, 13),
            text_color=TEXT_SECONDARY
        ).pack(anchor="w", pady=(0, 10))
        
        # Checkbox variables
        self.export_hourly_var = tkinter.BooleanVar(value=True)
        self.export_monthly_var = tkinter.BooleanVar(value=True)
        self.export_yearly_var = tkinter.BooleanVar(value=True)
        self.export_metadata_var = tkinter.BooleanVar(value=True)
        
        # Checkboxes
        customtkinter.CTkCheckBox(
            options_frame,
            text="Hourly time series data",
            variable=self.export_hourly_var,
            font=(FONT_FAMILY_TEXT, 12)
        ).pack(anchor="w", pady=5)
        
        customtkinter.CTkCheckBox(
            options_frame,
            text="Monthly aggregated summary",
            variable=self.export_monthly_var,
            font=(FONT_FAMILY_TEXT, 12)
        ).pack(anchor="w", pady=5)
        
        customtkinter.CTkCheckBox(
            options_frame,
            text="Yearly totals and KPIs",
            variable=self.export_yearly_var,
            font=(FONT_FAMILY_TEXT, 12)
        ).pack(anchor="w", pady=5)
        
        customtkinter.CTkCheckBox(
            options_frame,
            text="Import configuration & metadata",
            variable=self.export_metadata_var,
            font=(FONT_FAMILY_TEXT, 12)
        ).pack(anchor="w", pady=5)
        
        # Status Label
        self.export_tab_status_label = customtkinter.CTkLabel(
            scroll_container,
            text="Ready to export to local PC",
            font=(FONT_FAMILY_TEXT, 12),
            text_color=TEXT_SECONDARY
        )
        self.export_tab_status_label.pack(anchor="w", pady=(20, 30))
        
        # Buttons Frame
        buttons_frame = customtkinter.CTkFrame(scroll_container, fg_color="transparent")
        buttons_frame.pack(fill="x", pady=(10, 20))
        
        # Export to Excel (Blue, Primary)
        excel_btn = customtkinter.CTkButton(
            buttons_frame,
            text="Export to Excel (.xlsx)",
            command=self._export_to_excel_tab,
            fg_color=THEME.primary.blue_medium,
            hover_color=THEME.primary.blue_deep,
            text_color="white",
            width=180,
            height=40,
            corner_radius=8,
            font=(FONT_FAMILY_TEXT, 13, "bold")
        )
        excel_btn.pack(side="left", padx=(0, 10))
        
        # Export to CSV (Gray, Secondary)
        csv_btn = customtkinter.CTkButton(
            buttons_frame,
            text="Export to CSV (.csv)",
            command=self._export_to_csv_tab,
            fg_color=THEME.text.gray_medium,
            hover_color=THEME.text.gray_darker,
            text_color="white",
            width=180,
            height=40,
            corner_radius=8,
            font=(FONT_FAMILY_TEXT, 13)
        )
        csv_btn.pack(side="left", padx=(10, 0))
    def _prepare_hourly_dataframe(self, api_result):
        """
        Prepare data preserving EXACT API column structure.
        Different sources return different columns - preserve them all with original names.
        """
        raw_df = api_result["hourly_data"].copy()
        source = self.import_config.get("source", "UNKNOWN")
        mode = self.import_config.get("pvgis_mode", "HOURLY")
        
        
        time_col = None
        if 'time(UTC)' in raw_df.columns:
            time_col = 'time(UTC)'
            raw_df[time_col] = pd.to_datetime(raw_df[time_col], format="%Y%m%d:%H%M", utc=True)
        elif 'time' in raw_df.columns:
            time_col = 'time'
            if source == "NINJA":
                pass
            else:
                raw_df[time_col] = pd.to_datetime(raw_df[time_col], format="%Y%m%d:%H%M", utc=True)
        
        if time_col and time_col in raw_df.columns:
            raw_df.sort_values(time_col, inplace=True)
        
        if source == "PVGIS" and mode == "TMY":
            source_type = "PVGIS_TMY"
            is_tmy = True
            main_metric = "G(h)" if "G(h)" in raw_df.columns else None
        elif source == "PVGIS" and mode == "HOURLY":
            source_type = "PVGIS_HOURLY"
            is_tmy = False
            main_metric = "P" if "P" in raw_df.columns else None
        elif source == "NINJA":
            source_type = "NINJA"
            is_tmy = False
            main_metric = "electricity" if "electricity" in raw_df.columns else None
        else:
            source_type = "UNKNOWN"
            is_tmy = False
            main_metric = None
        
        return raw_df, source_type, is_tmy, main_metric

    def _compute_monthly_view(self, hourly_df):
        """
        Aggregate hourly to monthly.
        - Power/energy/irradiance columns: SUM (total energy)
        - Weather parameters (temp, wind): MEAN (average)
        """
        df = hourly_df.copy()
        
        datetime_col = None
        for col in df.columns:
            if 'time' in str(col).lower():
                datetime_col = col
                break
        
        if datetime_col is None:
            raise ValueError("No datetime column found in data")
        
        df[datetime_col] = pd.to_datetime(df[datetime_col])
        df.set_index(datetime_col, inplace=True)
        
        exclude_cols = ['Month', 'Year', 'Month Name']
        numeric_cols = [c for c in df.select_dtypes(include=[np.number]).columns 
                       if c not in exclude_cols]
        
        # Categorize columns for aggregation
        sum_cols = []   # Power, energy, irradiance (to be summed)
        mean_cols = []  # Weather parameters (to be averaged)
        
        for col in numeric_cols:
            col_str = str(col)
            col_lower = col_str.lower()
            
            # Power/energy columns (to sum)
            if col in ['P', 'electricity']:
                sum_cols.append(col)
            # Irradiance columns (Gb, Gd, Gr, G) - sum gives irradiation
            elif any(x in col_lower for x in ['gb', 'gd', 'gr', 'g(i)', 'g_h', 'g_d', 'g_b']):
                sum_cols.append(col)
            # Everything else (temperature, wind, sun angle, etc.) - average
            else:
                mean_cols.append(col)
        
        # Aggregate: sum for power/energy/irradiance, mean for weather
        monthly_parts = []
        counts = df.resample('ME').size()
        if sum_cols:
            monthly_parts.append(df.resample('ME')[sum_cols].sum())
        if mean_cols:
            monthly_parts.append(df.resample('ME')[mean_cols].mean())
        
        # Combine results
        if len(monthly_parts) > 1:
            monthly = pd.concat(monthly_parts, axis=1)
        elif monthly_parts:
            monthly = monthly_parts[0]
        else:
            monthly = pd.DataFrame()
            
        # Drop rows where there was absolutely no data in the original hourly dataframe
        if not monthly.empty:
            monthly = monthly[counts > 0]
        
        # Ensure consistent column order (only include cols that exist)
        valid_cols = [c for c in numeric_cols if c in monthly.columns]
        monthly = monthly[valid_cols]
        monthly.reset_index(inplace=True)
        
        monthly['Month'] = monthly[datetime_col].dt.month
        monthly['Year'] = monthly[datetime_col].dt.year
        monthly['Month Name'] = monthly['Month'].apply(lambda x: calendar.month_abbr[int(x)])
        
        # Rename power→energy only (P→E); keep irradiance names G*
        RENAME_AGGREGATED = {"P": "E"}
        rename_apply = {k: v for k, v in RENAME_AGGREGATED.items() if k in monthly.columns}
        monthly = monthly.rename(columns=rename_apply)
        final_cols = [datetime_col, 'Month Name', 'Year'] + [rename_apply.get(c, c) for c in valid_cols]
        monthly = monthly[final_cols]
        # Monthly: store in kWh, kWh/m² (÷1000)
        _energy_cols = [c for c in ('E', 'Esource', 'electricity') if c in monthly.columns]
        _irrad_cols = [c for c in ('G(i)', 'G(h)', 'Gb(i)', 'Gd(i)', 'Gr(i)') if c in monthly.columns]
        if _energy_cols:
            monthly[_energy_cols] = monthly[_energy_cols] / 1000.0
        if _irrad_cols:
            monthly[_irrad_cols] = monthly[_irrad_cols] / 1000.0
        return monthly

    def _compute_yearly_view(self, monthly_df):
        """
        Aggregate monthly to yearly.
        - Power/energy/irradiance columns: SUM (total energy)
        - Weather parameters (temp, wind): MEAN (average)
        """
        df = monthly_df.copy()
        
        exclude = ['Month', 'Month Name', 'Year']
        time_cols = []
        for col in df.columns:
            if 'time' in str(col).lower():
                exclude.append(col)
                time_cols.append(col)
        
        numeric_cols = [c for c in df.select_dtypes(include=[np.number]).columns 
                       if c not in exclude]
        
        # Categorize columns for aggregation (same as monthly; monthly has E and same G* names)
        sum_cols = []   # Power/energy, irradiance (to be summed)
        mean_cols = []  # Weather parameters (to be averaged)
        
        for col in numeric_cols:
            col_str = str(col)
            col_lower = col_str.lower()
            
            # Power/energy columns (to sum)
            if col in ['P', 'electricity', 'E', 'Esource']:
                sum_cols.append(col)
            # Irradiance columns (G* - same names in monthly/yearly)
            elif any(x in col_lower for x in ['gb', 'gd', 'gr', 'g(i)', 'g_h', 'g_d', 'g_b']):
                sum_cols.append(col)
            # Everything else (temperature, wind, etc.) - average
            else:
                mean_cols.append(col)
        
        # Aggregate by Year
        yearly_parts = []
        counts = df.groupby('Year').size()
        if sum_cols:
            yearly_parts.append(df.groupby('Year')[sum_cols].sum())
        if mean_cols:
            yearly_parts.append(df.groupby('Year')[mean_cols].mean())
        
        # Combine results
        if len(yearly_parts) > 1:
            yearly = pd.concat(yearly_parts, axis=1)
        elif yearly_parts:
            yearly = yearly_parts[0]
        else:
            yearly = pd.DataFrame()
            
        # Drop rows where there was absolutely no data
        if not yearly.empty:
            yearly = yearly[counts > 0]
        
        # Ensure consistent column order (only include cols that exist)
        valid_cols = [c for c in numeric_cols if c in yearly.columns]
        yearly = yearly[valid_cols]
        yearly.reset_index(inplace=True)
        # Yearly: monthly was in kWh/kWh/m²; sum = kWh per year → store in MWh, MWh/m² (÷1000)
        _energy_cols = [c for c in ('E', 'Esource', 'electricity') if c in yearly.columns]
        _irrad_cols = [c for c in ('G(i)', 'G(h)', 'Gb(i)', 'Gd(i)', 'Gr(i)') if c in yearly.columns]
        if _energy_cols:
            yearly[_energy_cols] = yearly[_energy_cols] / 1000.0
        if _irrad_cols:
            yearly[_irrad_cols] = yearly[_irrad_cols] / 1000.0
        return yearly

    def _ensure_view_data(self):
        """Prepare all views dynamically based on actual API response."""
        if hasattr(self, '_view_data_prepared') and self._view_data_prepared:
            return
        
        if not self.import_config or "api_result" not in self.import_config:
            return
        api_result = self.import_config["api_result"]
        
        hourly_df, source_type, is_tmy, main_metric = self._prepare_hourly_dataframe(api_result)
        
        monthly_df = self._compute_monthly_view(hourly_df)
        yearly_df = self._compute_yearly_view(monthly_df)
        
        self._view_hourly = hourly_df
        self._view_monthly = monthly_df
        self._view_yearly = yearly_df
        self._view_source_type = source_type
        self._view_is_tmy = is_tmy
        
        self._compute_kpis(hourly_df, source_type, is_tmy)
        
        
        self._view_data_prepared = True

    def _compute_kpis(self, hourly_df, source_type, is_tmy):
        """
        Compute KPIs from hourly data.
        Follows official source calculation methods.
        """
        config = self.import_config or {}
        
        if is_tmy:
            if 'G(h)' in hourly_df.columns:
                quantity_col = 'G(h)'
                total_quantity = hourly_df[quantity_col].sum() / 1000
                unit = "kWh/m²"
                data_type = "irradiation"
            else:
                total_quantity = 0
                unit = "kWh/m²"
                data_type = "irradiation"
        elif source_type == "NINJA":
            if 'electricity' in hourly_df.columns:
                quantity_col = 'electricity'
                total_quantity = hourly_df[quantity_col].sum()
                unit = "kWh"
                data_type = "energy"
            else:
                total_quantity = 0
                unit = "kWh"
                data_type = "energy"
        else:  # PVGIS_HOURLY
            if 'P' in hourly_df.columns:
                quantity_col = 'P'
                total_quantity = hourly_df[quantity_col].sum() / 1000
                unit = "kWh"
                data_type = "energy"
            else:
                total_quantity = 0
                unit = "kWh"
                data_type = "energy"
        
        if source_type == "PVGIS_HOURLY":
            capacity_kw = config.get('peak_power_kwp', 1.0)
        elif source_type == "NINJA":
            capacity_kw = config.get('capacity_kw', 1.0)
        else:
            capacity_kw = 1.0
        
        hours_in_year = len(hourly_df)
        
        annual_energy = total_quantity
        
        if capacity_kw > 0:
            specific_yield = annual_energy / capacity_kw
        else:
            specific_yield = 0
        
        if capacity_kw > 0 and hours_in_year > 0:
            capacity_factor = (annual_energy / (capacity_kw * hours_in_year)) * 100
        else:
            capacity_factor = 0
        
        self._kpi_annual_energy = annual_energy
        self._kpi_specific_yield = specific_yield
        self._kpi_capacity_factor = capacity_factor
        self._kpi_unit = unit
        self._kpi_data_type = data_type
        self._kpi_capacity_kw = capacity_kw
        self._kpi_hours = hours_in_year
        
    def _open_data_window(self, title, df, data_type):
        """Apple-style modal: frosted overlay, clean header, data badge."""
        try:
            if df is None or len(df) == 0:
                self._show_no_data_dialog(title)
                return
            
            config = self.import_config or {}
            lat = config.get('latitude', 'N/A')
            lon = config.get('longitude', 'N/A')

            # When opened from Visualize & Export popup, show overlay/table inside the popup
            modal_root = getattr(self, '_visualize_export_modal_root', None)
            if modal_root and modal_root.winfo_exists():
                parent_for_modal = modal_root
            else:
                parent_for_modal = self
            
            # Frosted overlay – medium gray simulates 50% dim
            overlay = customtkinter.CTkFrame(
                parent_for_modal, fg_color="#D1D1D6", corner_radius=0,
            )
            overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
            overlay.lift()
            # Click overlay to close
            overlay.bind("<Button-1>", lambda e: self._elegant_close(glass, overlay))
            
            # Modal card
            glass = customtkinter.CTkFrame(
                parent_for_modal, fg_color=BG_CARD, corner_radius=20,
                border_width=1, border_color=THEME.border.light,
            )
            glass.place(relx=0.5, rely=0.5, anchor="center",
                       relwidth=0.92, relheight=0.88)
            glass.lift()
            
            # Inner container
            inner = customtkinter.CTkFrame(glass, fg_color="transparent")
            inner.pack(fill="both", expand=True, padx=2, pady=2)
            
            # ── Header ──
            header = customtkinter.CTkFrame(inner, fg_color="transparent", height=60)
            header.pack(fill="x", padx=28, pady=(20, 0))
            header.pack_propagate(False)
            
            # Left: title + location
            left = customtkinter.CTkFrame(header, fg_color="transparent")
            left.pack(side="left")
            customtkinter.CTkLabel(
                left, text=title,
                font=(FONT_FAMILY_DISPLAY, 18, "bold"),
                text_color=THEME.text.primary,
            ).pack(anchor="w")
            customtkinter.CTkLabel(
                left, text=f"Lat {lat}, Lon {lon}",
                font=(FONT_FAMILY_TEXT, 12),
                text_color=THEME.text.muted,
            ).pack(anchor="w")
            
            # Right: Export + Close
            right = customtkinter.CTkFrame(header, fg_color="transparent")
            right.pack(side="right")
            
            customtkinter.CTkButton(
                right, text="⬇ Export", width=100, height=36,
                command=lambda: self._export_current_df(df, title),
                fg_color=THEME.primary.blue, hover_color=THEME.primary.blue_hover,
                text_color="#FFFFFF", corner_radius=10,
                font=(FONT_FAMILY_TEXT, 12, "bold"),
            ).pack(side="left", padx=(0, 10))
            
            # Circle close button
            customtkinter.CTkButton(
                right, text="✕", width=36, height=36,
                command=lambda: self._elegant_close(glass, overlay),
                fg_color=THEME.bg.hover, hover_color=THEME.border.light,
                text_color=THEME.text.secondary, corner_radius=18,
                font=(FONT_FAMILY_TEXT, 16),
            ).pack(side="left")
            
            # Divider
            customtkinter.CTkFrame(
                inner, fg_color=THEME.border.light, height=1, corner_radius=0,
            ).pack(fill="x", padx=24, pady=(12, 0))
            
            # Source badge
            badge_text = "ENERGY RESULTS" if data_type == "energy" else (getattr(self, "_view_source_type", "DATA").replace("_", " ").upper())
            badge_row = customtkinter.CTkFrame(inner, fg_color="transparent")
            badge_row.pack(fill="x", padx=28, pady=(12, 0))
            customtkinter.CTkLabel(
                badge_row,
                text=badge_text,
                font=(FONT_FAMILY_TEXT, 10, "bold"),
                text_color=THEME.primary.blue,
                fg_color=THEME.primary.blue_light,
                corner_radius=6, padx=10, pady=4,
            ).pack(side="left")
            
            # Table container – show first 3000 rows for speed; export still uses full data
            table_box = customtkinter.CTkFrame(
                inner, fg_color=THEME.bg.gray_pale, corner_radius=14,
            )
            table_box.pack(fill="both", expand=True, padx=20, pady=16)
            display_df = df.head(3000) if len(df) > 3000 else df
            if len(df) > 3000:
                trunc_label = customtkinter.CTkLabel(
                    table_box,
                    text=f"Showing first 3,000 of {fmt_num(len(df))} rows. Export to download all data.",
                    font=(FONT_FAMILY_TEXT, 11),
                    text_color=THEME.text.secondary,
                )
                trunc_label.pack(pady=(12, 4))
            # Render table (context=energy for energy results, else import; view for kWh/MWh units)
            table_context = "energy" if data_type == "energy" else "import"
            if data_type == "energy":
                view = "daily" if title == "Daily" else ("monthly" if title == "Monthly" else ("yearly" if title == "Yearly" else "hourly"))
            else:
                view = "hourly" if "hourly" in data_type else ("monthly" if "monthly" in data_type else "yearly")
            try:
                self._render_elegant_table(table_box, display_df, context=table_context, view=view)
            except Exception as e:
                customtkinter.CTkLabel(
                    table_box, text=f"Error: {str(e)}",
                    text_color=THEME.status.error_alt
                ).pack(expand=True)
            
            # Animate in
            self._elegant_animate_in(glass)
            
            # Close on overlay click
            overlay.bind("<Button-1>", lambda e: self._elegant_close(glass, overlay))
            glass.bind("<Escape>", lambda e: self._elegant_close(glass, overlay))
            
        except Exception as e:
            import traceback
            traceback.print_exc()

    def _elegant_animate_in(self, popup):
        """Subtle slide-up entrance."""
        popup.place_configure(rely=0.53)
        popup.after(30, lambda: popup.place_configure(rely=0.515))
        popup.after(60, lambda: popup.place_configure(rely=0.5))

    def _elegant_close(self, popup, overlay):
        """Subtle slide-down exit."""
        popup.place_configure(rely=0.515)
        popup.after(30, lambda: popup.place_configure(rely=0.53))
        popup.after(60, lambda: [popup.destroy(), overlay.destroy()])

    def _render_elegant_table(self, parent, df, context="import", view=None):
        """Render clean modern table. context: 'energy' | 'import'. view: 'hourly'|'monthly'|'yearly' for import G* unit (Wh/m² when aggregated)."""
        import tkinter.ttk as ttk
        
        style = ttk.Style()
        style.theme_use('clam')
        
        # Clean colors
        style.configure(
            "Elegant.Treeview",
            background=BG_CARD,
            foreground=THEME.text.slate_medium,
            fieldbackground=BG_CARD,
            rowheight=34,
            font=(FONT_FAMILY_TEXT, 11),
            borderwidth=0
        )
        style.configure(
            "Elegant.Treeview.Heading",
            background=THEME.bg.gray_pale,
            foreground=THEME.text.slate_light,
            font=(FONT_FAMILY_TEXT, 11, "bold"),
            relief="flat",
            padding=10
        )
        style.map(
            "Elegant.Treeview",
            background=[("selected", THEME.status.info_pale)],
            foreground=[("selected", PRIMARY_BLUE)]
        )
        
        # Container
        container = customtkinter.CTkFrame(parent, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=8, pady=8)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        
        columns = list(df.columns)
        tree = ttk.Treeview(
            container, columns=columns, show='headings',
            style="Elegant.Treeview"
        )
        
        # Scrollbars
        vsb = customtkinter.CTkScrollbar(
            container, orientation="vertical", command=tree.yview,
            fg_color=THEME.bg.divider, button_color=THEME.border.gray_dark
        )
        hsb = customtkinter.CTkScrollbar(
            container, orientation="horizontal", command=tree.xview,
            fg_color=THEME.bg.divider, button_color=THEME.border.gray_dark
        )
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        # Columns
        for col in columns:
            is_num = pd.api.types.is_numeric_dtype(df[col])
            is_dt = pd.api.types.is_datetime64_any_dtype(df[col])
            
            anchor = 'center' if is_dt else ('e' if is_num else 'w')
            width = 150 if is_dt else (100 if is_num else 130)
            
            unit = get_column_unit(col, context, view=view)
            display = format_column_header(str(col), unit)
            if len(display) > 20:
                display = display[:17] + "..."
            tree.heading(col, text=display, anchor='center')
            tree.column(col, width=width, anchor=anchor, stretch=True)
        
        # Data with alternating
        for idx, (_, row) in enumerate(df.iterrows()):
            vals = []
            for col, val in zip(columns, row):
                if pd.isna(val): vals.append("")
                elif str(col).lower() in ['year', 'month']:
                    try:
                        vals.append(str(int(val)))
                    except (ValueError, TypeError):
                        vals.append(str(val))
                elif pd.api.types.is_datetime64_any_dtype(df[col]):
                    vals.append(pd.Timestamp(val).strftime("%Y-%m-%d %H:%M"))
                elif isinstance(val, (int, np.integer)):
                    vals.append(fmt_num(int(val)))
                elif isinstance(val, (float, np.floating)):
                    vals.append(fmt_num(val, 3))
                else: vals.append(str(val))
            
            tag = 'alt' if idx % 2 else ''
            tree.insert('', 'end', values=tuple(vals), tags=(tag,))
        
        tree.tag_configure('alt', background=BG_SIDEBAR)

    def _show_no_data_dialog(self, title):
        """Show a simple 'no data' dialog."""
        no_data_win = customtkinter.CTkToplevel(self)
        no_data_win.title(title)
        no_data_win.geometry("400x200")
        no_data_win.transient(self)
        
        customtkinter.CTkLabel(
            no_data_win,
            text=f"No {title.lower()} available.",
            font=(FONT_FAMILY_TEXT, 13),
            text_color=TEXT_PRIMARY
        ).pack(expand=True)
        
        customtkinter.CTkButton(
            no_data_win,
            text="Close",
            command=no_data_win.destroy,
            fg_color=PRIMARY_BLUE,
            hover_color=PRIMARY_HOVER
        ).pack(pady=10)
        no_data_win.grab_set()

    def _show_error_dialog(self, message):
        """Show error dialog instead of crashing."""
        error_window = customtkinter.CTkToplevel(self)
        error_window.title("Error")
        error_window.geometry("400x200")
        error_window.transient(self)
        
        customtkinter.CTkLabel(
            error_window,
            text="⚠️ Error",
            font=(FONT_FAMILY_TEXT, 16, "bold"),
            text_color=ERROR_RED
        ).pack(pady=(20, 10))
        
        customtkinter.CTkLabel(
            error_window,
            text=message,
            font=(FONT_FAMILY_TEXT, 12),
            wraplength=350
        ).pack(pady=10)
        
        customtkinter.CTkButton(
            error_window,
            text="OK",
            command=error_window.destroy
        ).pack(pady=20)

    # ============================================
    # EXPORT METHODS (moved from Step 3)
    # ============================================
    
    def _export_to_excel(self):
        """Export data to Excel with multiple sheets"""
        status_lbl = getattr(self, 'export_tab_status_label', None) or getattr(self, 'export_status_label', None)
        if not EXCEL_AVAILABLE:
            if status_lbl and status_lbl.winfo_exists():
                status_lbl.configure(text="❌ openpyxl not installed. Run: pip install openpyxl", text_color=ERROR_RED)
            return
        
        try:
            # Check if data exists
            if not hasattr(self, '_view_hourly') or self._view_hourly is None:
                if status_lbl and status_lbl.winfo_exists():
                    status_lbl.configure(text="❌ No data to export. Complete Step 1 first.", text_color=ERROR_RED)
                return
            
            # Get filename from user
            config = self.import_config or {}
            lat = config.get('latitude', '0')
            lon = config.get('longitude', '0')
            default_name = f"PV_Data_{lat}_{lon}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_name
            )
            
            if not file_path:
                return
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Add sheets with data
            self._write_dataframe_to_sheet(wb, "Hourly Data", self._view_hourly)
            self._write_dataframe_to_sheet(wb, "Monthly Summary", self._view_monthly)
            self._write_dataframe_to_sheet(wb, "Yearly Summary", self._view_yearly)
            self._write_metadata_sheet(wb, "Metadata")
            
            # Save
            wb.save(file_path)
            
            if status_lbl and status_lbl.winfo_exists():
                status_lbl.configure(text=f"✓ Exported to: {file_path.split('/')[-1]}", text_color=SUCCESS_GREEN)
            
        except Exception as e:
            if status_lbl and status_lbl.winfo_exists():
                status_lbl.configure(text=f"❌ Export failed: {str(e)}", text_color=ERROR_RED)
    
    def _write_dataframe_to_sheet(self, workbook, sheet_name, df):
        """Write a DataFrame to a formatted Excel sheet"""
        if df is None or df.empty:
            ws = workbook.create_sheet(title=sheet_name)
            ws["A1"] = "No data available"
            return
        
        ws = workbook.create_sheet(title=sheet_name)
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Format value based on type
                if pd.isna(value):
                    cell.value = ""
                elif isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00'
                elif isinstance(value, datetime):
                    cell.value = value
                    cell.number_format = 'YYYY-MM-DD HH:MM'
                else:
                    cell.value = str(value)
        
        # Auto-adjust column widths
        for col_idx in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _write_metadata_sheet(self, workbook, sheet_name):
        """Write metadata/configuration info to a sheet"""
        ws = workbook.create_sheet(title=sheet_name)
        
        config = self.import_config or {}
        
        # Build metadata based on data type
        metadata = [
            ["Export Metadata", ""],
            ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Source", config.get("source", "N/A")],
        ]
        
        if config.get("source") == "NINJA":
            metadata.append(["Mode", "PV"])
        else:
            metadata.append(["Mode", config.get("pvgis_mode", "N/A")])
        
        metadata.append(["Location (Lat, Lon)", f"{config.get('latitude', 'N/A')}, {config.get('longitude', 'N/A')}"])
        
        if config.get("source") == "NINJA":
            # Ninja PV-specific metadata
            metadata.append(["Dataset", config.get("dataset", "N/A")])
            metadata.append(["Year", config.get("year", "N/A")])
            metadata.append(["Capacity (kW)", config.get("capacity_kw", "N/A")])
            metadata.append(["PV height (m)", config.get("pv_height_m", "N/A")])
            metadata.append(["System Loss (%)", f"{config.get('system_loss_fraction', 0) * 100:.1f}"])
            metadata.append(["Tracking", config.get("tracking", "N/A")])
            metadata.append(["Tilt", config.get("tilt", "N/A")])
            metadata.append(["Azimuth", config.get("azimuth", "N/A")])
        else:
            # PVGIS metadata
            metadata.append(["Database", config.get("database", config.get("tmy_database", "N/A"))])
            metadata.append(["Year Range", f"{config.get('start_year', 'N/A')} - {config.get('end_year', 'N/A')}"])
            metadata.append(["Peak Power (kWp)", config.get("peak_power_kwp", "N/A")])
            metadata.append(["System Loss (%)", config.get("system_loss_percent", "N/A")])
            metadata.append(["Mounting Type", config.get("mounting_type", "N/A")])
            metadata.append(["PV Technology", config.get("pv_technology", "N/A")])
        
        metadata.extend([
            ["", ""],
            ["KPIs", ""],
            ["Annual Energy", f"{fmt_num(getattr(self, '_kpi_annual_energy', 0), 2)} {getattr(self, '_kpi_unit', 'kWh')}"],
            ["Specific Yield", fmt_num(getattr(self, '_kpi_specific_yield', 0), 2)],
            ["Capacity Factor", f"{getattr(self, '_kpi_capacity_factor', 0):.2f}%"],
        ])
        
        for row_idx, (key, value) in enumerate(metadata, 1):
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40
    
    def _export_to_csv(self):
        """Export data to CSV (placeholder for now)"""
        status_lbl = getattr(self, 'export_tab_status_label', None) or getattr(self, 'export_status_label', None)
        if status_lbl and status_lbl.winfo_exists():
            status_lbl.configure(text="CSV export - coming soon", text_color=TEXT_SECONDARY)
    
    def _export_to_excel_tab(self):
        """Export selected data to Excel file on local PC from Export tab"""
        # Validation
        if not hasattr(self, '_view_hourly') or self._view_hourly is None or len(self._view_hourly) == 0:
            self.export_tab_status_label.configure(
                text="Error: No data. Complete Step 1 first.",
                text_color=ERROR_RED
            )
            return
        
        if not any([self.export_hourly_var.get(), self.export_monthly_var.get(),
                    self.export_yearly_var.get(), self.export_metadata_var.get()]):
            self.export_tab_status_label.configure(
                text="Error: Select at least one data type.",
                text_color=ERROR_RED
            )
            return
        
        if not EXCEL_AVAILABLE:
            self.export_tab_status_label.configure(
                text="Error: openpyxl not installed. Run: pip install openpyxl",
                text_color=ERROR_RED
            )
            return
        
        # Generate filename
        config = self.import_config or {}
        lat = config.get('latitude', '0.0')
        lon = config.get('longitude', '0.0')
        db = config.get('database', 'PVGIS')
        if db:
            db = db.replace(' ', '_')
        else:
            db = config.get('tmy_database', 'PVGIS').replace(' ', '_')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"PVGIS_{lat}_{lon}_{db}_{timestamp}.xlsx"
        
        # File dialog - SAVES TO LOCAL PC
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Excel File to PC"
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # Hourly Sheet
            if self.export_hourly_var.get() and hasattr(self, '_view_hourly') and self._view_hourly is not None:
                ws = wb.create_sheet("Hourly Data")
                self._write_df_to_sheet(ws, self._view_hourly, freeze=True)
            
            # Monthly Sheet
            if self.export_monthly_var.get() and hasattr(self, '_view_monthly') and self._view_monthly is not None:
                ws = wb.create_sheet("Monthly Summary")
                self._write_df_to_sheet(ws, self._view_monthly)
            
            # Yearly Sheet
            if self.export_yearly_var.get() and hasattr(self, '_view_yearly') and self._view_yearly is not None:
                ws = wb.create_sheet("Yearly Summary")
                self._write_df_to_sheet(ws, self._view_yearly)
            
            # Metadata Sheet
            if self.export_metadata_var.get():
                ws = wb.create_sheet("Metadata")
                self._write_metadata_to_sheet(ws)
            
            wb.save(file_path)
            
            self.export_tab_status_label.configure(
                text=f"✓ Saved to PC: {file_path.split('/')[-1]}",
                text_color=SUCCESS_GREEN
            )
            
        except Exception as e:
            self.export_tab_status_label.configure(
                text=f"Export failed: {str(e)}",
                text_color=ERROR_RED
            )
            show_toast(self, str(e, type="error"))
    
    def _export_to_csv_tab(self):
        """Export selected data to MULTIPLE CSV files (one per checked data type)"""
        if not hasattr(self, '_view_hourly') or self._view_hourly is None:
            self.export_tab_status_label.configure(
                text="Error: No data to export.",
                text_color=ERROR_RED
            )
            return
        
        if not any([self.export_hourly_var.get(), self.export_monthly_var.get(),
                    self.export_yearly_var.get(), self.export_metadata_var.get()]):
            self.export_tab_status_label.configure(
                text="Error: Select at least one data type.",
                text_color=ERROR_RED
            )
            return
        
        config = self.import_config or {}
        lat = config.get('latitude', '0.0')
        lon = config.get('longitude', '0.0')
        db = config.get('database', 'PVGIS')
        if db:
            db = db.replace(' ', '_')
        else:
            db = config.get('tmy_database', 'PVGIS').replace(' ', '_')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Ask user to select folder
        folder_path = filedialog.askdirectory(title="Select Folder to Save CSV Files")
        
        if not folder_path:
            return
        
        try:
            import os
            
            # Helper to strip timezone
            def strip_tz(df):
                df_copy = df.copy()
                for col in df_copy.columns:
                    if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                        if df_copy[col].dt.tz is not None:
                            df_copy[col] = df_copy[col].dt.tz_localize(None)
                return df_copy
            
            base_name = f"PVGIS_{lat}_{lon}_{db}_{timestamp}"
            exported_files = []
            total_rows = 0
            
            # 1. Hourly Data
            if self.export_hourly_var.get() and hasattr(self, '_view_hourly') and self._view_hourly is not None:
                hourly_path = os.path.join(folder_path, f"{base_name}_Hourly_Data.csv")
                hourly_df = strip_tz(self._view_hourly)
                hourly_df.to_csv(hourly_path, index=False, encoding='utf-8')
                exported_files.append("Hourly Data")
                total_rows += len(hourly_df)
            
            # 2. Monthly Data
            if self.export_monthly_var.get() and hasattr(self, '_view_monthly') and self._view_monthly is not None:
                monthly_path = os.path.join(folder_path, f"{base_name}_Monthly_Summary.csv")
                monthly_df = strip_tz(self._view_monthly)
                monthly_df.to_csv(monthly_path, index=False, encoding='utf-8')
                exported_files.append("Monthly Summary")
                total_rows += len(monthly_df)
            
            # 3. Yearly Data
            if self.export_yearly_var.get() and hasattr(self, '_view_yearly') and self._view_yearly is not None:
                yearly_path = os.path.join(folder_path, f"{base_name}_Yearly_Summary.csv")
                yearly_df = strip_tz(self._view_yearly)
                yearly_df.to_csv(yearly_path, index=False, encoding='utf-8')
                exported_files.append("Yearly Summary")
                total_rows += len(yearly_df)
            
            # 4. Metadata
            if self.export_metadata_var.get():
                cfg = self.import_config or {}
                if cfg.get('source') == "PVGIS" and cfg.get('pvgis_mode') == "TMY":
                    year_text = f"TMY - {cfg.get('tmy_database', 'N/A')}"
                elif cfg.get('source') == "PVGIS":
                    year_text = f"{cfg.get('start_year', 'N/A')} - {cfg.get('end_year', 'N/A')}"
                else:
                    year_text = str(cfg.get('year', 'N/A'))
                
                metadata_path = os.path.join(folder_path, f"{base_name}_Metadata.csv")
                metadata_data = [
                    ["PV Desalination System - Import Configuration", ""],
                    ["", ""],
                    ["Parameter", "Value"],
                    ["Source Type", cfg.get('source', 'N/A')],
                    ["Mode", cfg.get('pvgis_mode', 'N/A')],
                    ["Latitude", cfg.get('latitude', 'N/A')],
                    ["Longitude", cfg.get('longitude', 'N/A')],
                    ["Database", cfg.get('database', cfg.get('tmy_database', 'N/A'))],
                    ["Time Period", year_text],
                    ["Peak Power (kWp)", cfg.get('peak_power_kwp', cfg.get('capacity_kw', 'N/A'))],
                    ["System Loss (%)", cfg.get('system_loss_percent', cfg.get('system_loss_fraction', 'N/A'))],
                    ["Mounting Type", cfg.get('mounting_type', cfg.get('tracking', 'N/A'))],
                    ["Tilt Angle (°)", cfg.get('slope', cfg.get('tilt', 'N/A'))],
                    ["Azimuth (°)", cfg.get('azimuth', cfg.get('aspect', 'N/A'))],
                    ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M")],
                    ["", ""],
                    ["KPIs", ""],
                    ["Annual Energy", f"{fmt_num(getattr(self, '_kpi_annual_energy', 0), 2)} {getattr(self, '_kpi_unit', 'kWh')}"],
                    ["Specific Yield", fmt_num(getattr(self, '_kpi_specific_yield', 0), 2)],
                    ["Capacity Factor", f"{getattr(self, '_kpi_capacity_factor', 0):.2f}%"],
                ]
                metadata_df = pd.DataFrame(metadata_data)
                metadata_df.to_csv(metadata_path, index=False, header=False, encoding='utf-8')
                exported_files.append("Metadata")
            
            files_str = ", ".join(exported_files)
            self.export_tab_status_label.configure(
                text=f"✓ Saved {len(exported_files)} files: {files_str} ({total_rows} data rows)",
                text_color=SUCCESS_GREEN
            )
            
        except Exception as e:
            self.export_tab_status_label.configure(
                text=f"CSV export failed: {str(e)}",
                text_color=ERROR_RED
            )
    
    def _write_df_to_sheet(self, worksheet, df, freeze=False):
        """Write DataFrame to Excel sheet with formatting"""
        if df is None or len(df) == 0:
            worksheet.append(["No data"])
            return
        
        # FIX: Convert timezone-aware datetimes to naive (remove tzinfo)
        df_copy = df.copy()
        for col in df_copy.columns:
            if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                # Remove timezone info if present
                if df_copy[col].dt.tz is not None:
                    df_copy[col] = df_copy[col].dt.tz_localize(None)
        
        # Headers
        headers = list(df_copy.columns)
        worksheet.append(headers)
        
        # Format headers
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row in dataframe_to_rows(df_copy, index=False, header=False):
            worksheet.append(row)
        
        # Format numbers
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        
        # Auto-fit columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Freeze header
        if freeze:
            worksheet.freeze_panes = "A2"
    
    def _write_metadata_to_sheet(self, worksheet):
        """Write metadata to Excel sheet"""
        config = self.import_config or {}
        
        worksheet.append(["PV Desalination System - Import Configuration"])
        worksheet['A1'].font = Font(bold=True, size=14, color="3B82F6")
        worksheet.merge_cells('A1:B1')
        worksheet.append([])
        
        # Determine year range text
        if config.get('source') == "PVGIS" and config.get('pvgis_mode') == "TMY":
            year_text = f"TMY - {config.get('tmy_database', 'N/A')}"
        elif config.get('source') == "PVGIS":
            year_text = f"{config.get('start_year', 'N/A')} - {config.get('end_year', 'N/A')}"
        else:
            year_text = str(config.get('year', 'N/A'))
        
        metadata = [
            ("Parameter", "Value"),
            ("Source Type", config.get('source', 'N/A')),
            ("Mode", config.get('pvgis_mode', 'N/A')),
            ("Latitude", config.get('latitude', 'N/A')),
            ("Longitude", config.get('longitude', 'N/A')),
            ("Database", config.get('database', config.get('tmy_database', 'N/A'))),
            ("Time Period", year_text),
            ("Peak Power (kWp)", config.get('peak_power_kwp', config.get('capacity_kw', 'N/A'))),
            ("System Loss (%)", config.get('system_loss_percent', config.get('system_loss_fraction', 'N/A'))),
            ("Mounting Type", config.get('mounting_type', config.get('tracking', 'N/A'))),
            ("Tilt Angle (°)", config.get('slope', config.get('tilt', 'N/A'))),
            ("Azimuth (°)", config.get('azimuth', config.get('aspect', 'N/A'))),
            ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("", ""),
            ("KPIs", ""),
            ("Annual Energy", f"{fmt_num(getattr(self, '_kpi_annual_energy', 0), 2)} {getattr(self, '_kpi_unit', 'kWh')}"),
            ("Specific Yield", fmt_num(getattr(self, '_kpi_specific_yield', 0), 2)),
            ("Capacity Factor", f"{getattr(self, '_kpi_capacity_factor', 0):.2f}%"),
        ]
        
        for row_idx, (param, value) in enumerate(metadata, 3):
            worksheet.append([param, value])
            if param in ["Parameter", "Value", "KPIs"]:
                worksheet.cell(row=row_idx, column=1).font = Font(bold=True)
                if param != "":
                    worksheet.cell(row=row_idx, column=2).font = Font(bold=True)
            else:
                worksheet.cell(row=row_idx, column=1).font = Font(bold=True, color="6B7280")
        
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 40

    # ============================================
    # FINISH WORKFLOW METHOD
    # ============================================
    
    def _finish_source_workflow(self):
        """
        Complete the Source workflow from Step 2
        Stores data if checkbox checked, marks section complete, returns to home
        """
        try:
            # Store data if checkbox is checked
            if self.store_in_project_var.get():
                if hasattr(self, '_view_hourly') and self._view_hourly is not None:
                    self.app.source_data = {
                        'hourly': self._view_hourly,
                        'monthly': self._view_monthly if hasattr(self, '_view_monthly') else None,
                        'yearly': self._view_yearly if hasattr(self, '_view_yearly') else None,
                        'config': self.import_config,
                        'timestamp': datetime.now().isoformat()
                    }
                    print("Source data stored in app.source_data for later modules")
                else:
                    print("No data available to store – completing workflow anyway")
            
            # Mark section complete
            self.app.store.complete_section("source")
            
            # Return to home
            self.app.show_home_page()
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"ERROR in _finish_source_workflow: {e}")
            # Even on error, try to navigate home
            try:
                self.app.show_home_page()
            except Exception:
                pass
